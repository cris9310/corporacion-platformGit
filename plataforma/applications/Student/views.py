

from .models import *
from .forms import *
from applications.Finance.models import *
from applications.Programs.models import *
from applications.User.mixins import *


from django.views.generic import (TemplateView,
                                  FormView, CreateView, DeleteView, UpdateView,
                                  DetailView, ListView, View)
from django.urls import reverse_lazy, reverse
from django.http import JsonResponse
from django.shortcuts import render, HttpResponseRedirect, HttpResponse
from django.db.models import Avg
from django.db.models.base import Model as Model
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_control


import warnings
import pandas as pd
import re
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import Workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta




# Vista ok
@method_decorator(cache_control(no_cache=True, must_revalidate=True, no_store=True), name='dispatch')
class StudentTemplateView(AdminRequiredMixin,TemplateView): 
    template_name = "estudiantes/home_student.html"



# Vista que sirve para la creación de estudiantes de manera individual, se encuentra ok
class StudentCreateView(AdminRequiredMixin,CreateView):
    model = Estudiante
    form_class = StudentRegisterForm
    template_name = 'estudiantes/register_student.html'
    success_url = reverse_lazy('student_app:list-student')

    def post(self, request, *args, **kwargs):

        form = StudentRegisterForm(self.request.POST) 
        if form.is_valid():
            inv_list = []
            # Generador de código para un nuevo estudiante
            codigo = User.objects.code_generator()
            # Creando estudiantes en la bd de estudiante
            create_student = Estudiante.objects.create(
                codigo=codigo,
                tDocument=CatalogsTypesDocuement.objects.get(
                    nombre=form.cleaned_data.get("tDocument")),
                apellidos=form.cleaned_data.get('apellidos'),
                username=form.cleaned_data.get('username'),
                direccion=form.cleaned_data.get('direccion'),
                nacimiento=form.cleaned_data.get('nacimiento'),
                sexo=form.cleaned_data.get('sexo'),
                telefono=form.cleaned_data.get('telefono'),
                email=form.cleaned_data.get('email'),
                cedula=form.cleaned_data.get('cedula'),
                nombre=form.cleaned_data.get('nombre'),
                nacionalidad=form.cleaned_data.get('nacionalidad'),
                carrera=Programas.objects.get(
                    programa_name=form.cleaned_data.get('carrera')),
                sede=CatalogsSede.objects.get(
                    sede=form.cleaned_data.get('sede')),
                costo_cierre=Programas.objects.get(
                    programa_name=form.cleaned_data.get('carrera')).costo,
                nombre_acudiente=form.cleaned_data.get('nombre_acudiente'),
                apellidos_acudiente=form.cleaned_data.get(
                    'apellidos_acudiente'),
                telefono_acudiente=form.cleaned_data.get('telefono_acudiente'),
                cedula_acudiente=form.cleaned_data.get('cedula_acudiente'),
                periodo_matriculado=Periodos.objects.get(
                    periodo=form.cleaned_data.get('periodo_matriculado')),
                document=form.cleaned_data.get('document'),
                simat=form.cleaned_data.get('simat'),
                siet=form.cleaned_data.get('siet'),
                actaBachillerato=form.cleaned_data.get('actaBachillerato'),
                fotos=form.cleaned_data.get('fotos'),
                serviciosPublicos=form.cleaned_data.get('serviciosPublicos'),
                carneSalud=form.cleaned_data.get('carneSalud'),
                cedulaAcudiente=form.cleaned_data.get('cedulaAcudiente'),
                certificados=form.cleaned_data.get('certificados'),
                homologacion=form.cleaned_data.get('homologacion'),
                observaciones=form.cleaned_data.get('observaciones')
                                               
            )
            # Creamos usuario en la BD
            crea_user = User.objects.create_user(
                tipe=CatalogsTypesRol.objects.get(rol="Estudiante"),
                codigo=codigo,
                username=form.cleaned_data.get('username'),
                email=form.cleaned_data.get('email'),
                password=Estudiante.objects.get_secret("RANDOM"),
                nombres=form.cleaned_data.get('nombre'),
                apellidos=form.cleaned_data.get('apellidos'),
                is_superuser=False,
                is_active=True,
                is_staff=False,

            )
            # Generamos las facturas relacionadas con el estudiante, hace referencia a matrícula y a las facturas de las pensiones mensuales

            # vamos a realizar una validacion extra, verificamos que el programa al cual se matricula tenga o no grado y asi decidimos si 
            # generamos o no factura de derecho a grado. A cursos y diplomados no se genera facturas de matriculas
            # Recordar incluir las fechas de vencimiento de las facturas y analizar el hecho de incluir botones de cerrar o graduar a estudiante
            # si ya termina un programa que no tiene grado

            programa = Programas.objects.get(programa_name=form.cleaned_data.get('carrera'))
            cuotas = int(programa.cuotas)

            if programa.tiene_grado:
                # Calcular la fecha de vencimiento para el certificado de grado
                fechafacturacertificados = datetime.now() + relativedelta(months=cuotas)
                fechafacturacertificados = fechafacturacertificados.replace(day=5)

                # Crear factura de matrícula
                Facturas.objects.create(
                    user=User.objects.get(username=form.cleaned_data.get('username')),
                    codigo=Facturas.objects.code_invoice(),
                    email=form.cleaned_data.get('email'),
                    monto=programa.matricula,
                    due_at= datetime.now(),
                    descripcion="Matrícula",
                    estado=CatalogsTypesInvoices.objects.get(estado="Pendiente"),
                )

                # Crear facturas de mensualidades
                for i in range(cuotas):
                    if i == 0:
                        fechafacturamensualidades = datetime.now()
                    else:
                        fechafacturamensualidades = datetime.now() + relativedelta(months=i)
                        fechafacturamensualidades = fechafacturamensualidades.replace(day=5)

                    inv_list.append(Facturas(
                        user=User.objects.get(username=form.cleaned_data.get('username')),
                        codigo=int(Facturas.objects.code_invoice()) + i,
                        email=form.cleaned_data.get('email'),
                        monto=programa.cuota_valor,
                        due_at=fechafacturamensualidades,
                        descripcion="Mensualidad número " + str(i + 1),
                        estado=CatalogsTypesInvoices.objects.get(estado="Pendiente"),
                    ))

                # Crear todas las facturas de mensualidades en bulk
                Facturas.objects.bulk_create(inv_list)

                # Crear factura de derechos de grado
                Facturas.objects.create(
                    user=User.objects.get(username=form.cleaned_data.get('username')),
                    codigo=Facturas.objects.code_invoice(),
                    email=form.cleaned_data.get('email'),
                    monto=programa.derechosGrado,
                    due_at=fechafacturacertificados,
                    descripcion="Derechos de Grado",
                    estado=CatalogsTypesInvoices.objects.get(estado="Pendiente"),
                )

            else:
                # Calcular la fecha de vencimiento para el certificado de grado
                fechafacturacertificados = datetime.now() + relativedelta(months=cuotas)
                fechafacturacertificados = fechafacturacertificados.replace(day=5)

                # Crear factura de inscripción
                Facturas.objects.create(
                    user=User.objects.get(username=form.cleaned_data.get('username')),
                    codigo=Facturas.objects.code_invoice(),
                    email=form.cleaned_data.get('email'),
                    monto=programa.matricula,
                    due_at= datetime.now(),
                    descripcion="Inscripción",
                    estado=CatalogsTypesInvoices.objects.get(estado="Pendiente"),
                )

                # Crear facturas de mensualidades
                for i in range(cuotas):
                    if i == 0:
                        fechafacturamensualidades = datetime.now()
                    else:
                        fechafacturamensualidades = datetime.now() + relativedelta(months=i)
                        fechafacturamensualidades = fechafacturamensualidades.replace(day=5)
                    
                    inv_list.append(Facturas(
                        user=User.objects.get(username=form.cleaned_data.get('username')),
                        codigo=int(Facturas.objects.code_invoice()) + i,
                        email=form.cleaned_data.get('email'),
                        monto=programa.cuota_valor,
                        due_at=fechafacturamensualidades,
                        descripcion="Mensualidad número " + str(i + 1),
                        estado=CatalogsTypesInvoices.objects.get(estado="Pendiente"),
                    ))

                # Crear todas las facturas de mensualidades en bulk
                Facturas.objects.bulk_create(inv_list)

                # Crear factura de certificado
                Facturas.objects.create(
                    user=User.objects.get(username=form.cleaned_data.get('username')),
                    codigo=Facturas.objects.code_invoice(),
                    email=form.cleaned_data.get('email'),
                    monto=programa.derechosGrado,
                    due_at=fechafacturacertificados,
                    descripcion="Certificado",
                    estado=CatalogsTypesInvoices.objects.get(estado="Pendiente"),
                )

            mensaje = f'{self.model.__name__} registrado correctamente'
            error = "No hay error!"
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido registrar'
            error = form.errors
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 400
            return response


# Esta vista lista a todos los estudiantes que se han creado, se encuentra ok.
@method_decorator(cache_control(no_cache=True, must_revalidate=True, no_store=True), name='dispatch')
class Studentlistview(AdminRequiredMixin,ListView):
    model = Estudiante
    template_name = 'estudiantes/list_student.html'
    context_object_name = 'student'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['programas'] = Programas.objects.all()
        context['masivo'] = Estudiante.objects.filter(masivo=True).count()
        
        return context

    def get_queryset(self):

        data_student = []
        data_prin = Estudiante.objects.filter(is_graduado=False).exclude(masivo=True)
        for i in data_prin:
            if Banner.objects.filter(student__slug=i.slug).exists():
                data_json = {"pk": i.id, "slug":i.slug, "codigo": i.codigo, "nombres": i.nombre,
                                "apellidos": i.apellidos, "estado": True, "carrera": i.carrera,
                                "is_active": i.is_active, "is_matriculado": i.is_matriculado}
                data_student.append(data_json)
            else:
                data_json = {"pk": i.id, "slug":i.slug, "codigo": i.codigo, "nombres": i.nombre,
                                "apellidos": i.apellidos, "estado": False, "carrera": i.carrera,
                                "is_active": i.is_active, "is_matriculado": i.is_matriculado}
                data_student.append(data_json)
        queryset = data_student
        return queryset
    

# Vista para cargar estudiantes de forma masiva se encuentra ok

class StudentAsigView(AdminRequiredMixin,View):
    
    #momstramos el form que hemos creado
    def get(self, request, *args, **kwargs):
        form = StudentAsigMate()
        context = {"form": form}
        return render(request, r"estudiantes/asig_topics.html", context)

    def post(self, request, *args, **kwargs):

        # Declarando variables
        conteo = 0
        fecha_actual = date.today()

        # Declarando listas
        mensaje1 = []
        data_list = []
        l = ["Matricula", "Derechos de grado"]

        # Leyendo el archivo que recibimos desde el form
        new_studentData = self.request.FILES['carga']

        # zona de verificación de datos en el archivo
       
        if new_studentData.name == 'cargue_estudiantes.xlsx':
            warnings.filterwarnings(
                'ignore', category=UserWarning, module='openpyxl')
            new_student = pd.read_excel(
                new_studentData, sheet_name="Plantilla", engine='openpyxl')
            new_student['NACIMIENTO'] = pd.to_datetime(
                new_student['NACIMIENTO'])
            new_student['DUPLICADO'] = new_student.duplicated()
            new_student['DUPLICADO_USERS']= new_student['USERNAME'].duplicated()

            # Validamos que no existan valores nulos y en blancos en el archivo.
            miss_values_count = new_student.isnull().sum()
            miss_values_count = miss_values_count[miss_values_count != 0]

            # Generamos código para el estudiante
            codigo = User.objects.code_generator()

            # comparando los encabezados para ver si está correcta la estructura
            for i, a in zip(list(new_student), EST_STUDENT):
                if i != a:
                    break

            if miss_values_count.shape[0] or i != a:
                if miss_values_count.shape[0]:
                    if miss_values_count.shape[0] == 1:
                        mensaje1.append({"error": 'El archivo tiene ' + str(
                            miss_values_count.shape[0]) + ' columna sin datos, por favor, revise'})
                        response = JsonResponse(mensaje1, safe=False)
                        response.status_code = 400
                        return response
                    else:
                        mensaje1.append({"error": 'El archivo tiene ' + str(
                            miss_values_count.shape[0]) + ' columnas sin datos, por favor, revise'})
                        response = JsonResponse(mensaje1, safe=False)
                        response.status_code = 400
                        return response

                else:
                    mensaje1.append({"error": 'La columna número ' + str(i) +
                                    ' en su archivo, no tiene un encabezado válido, verifique el archivo y vuelva a cargarlo'})
                    response = JsonResponse(mensaje1, safe=False)
                    response.status_code = 400
                    return response
                    

            elif len(new_student) > 35 or len(new_student) == 0:
                if len(new_student) > 35:
                    mensaje1.append(
                        {"error": 'Recuerde que no puede cargar más de 35 estudiantes a la vez, en este archivo encontramos: ' + str(len(new_student))})
                    response = JsonResponse(mensaje1, safe=False)
                    response.status_code = 400
                    return response
                else:
                    mensaje1.append(
                        {"error": 'No encontramos estudiantes para cargar'})
                    response = JsonResponse(mensaje1, safe=False)
                    response.status_code = 400
                    return response

            else:
                # Validamos los campos

                for i in range(len(new_student)):
                    if new_student['DUPLICADO'][i] == True:
                        conteo = 1 + conteo
                        mensaje1.append({"error": "Encontramos filas duplicadas para el usuario: " +
                                        new_student['USERNAME'][i] + " verifique la información"})
                    else:
                        conteo = 0 + conteo

                for i in range(len(new_student)):
                    if new_student['DUPLICADO_USERS'][i] == True:
                        conteo = 1 + conteo
                        mensaje1.append({"error": "El usuario: " +
                                        new_student['USERNAME'][i] + " se encuentra duplicado, por favor verifique la información"})
                    else:
                        conteo = 0 + conteo

                for i in range(len(new_student)):
                    if User.objects.filter(username=new_student['USERNAME'][i]):
                        conteo = 1 + conteo
                        mensaje1.append(
                            {"error": "El nombre de usuario " + new_student['USERNAME'][i] + " ya existe, por favor cambie este username y vuelva a intentarlo."})
                    else:
                        conteo = 0 + conteo

                for i in range(len(new_student)):

                    res = " " in new_student['USERNAME'][i]
                    if res:
                        conteo = 1 + conteo
                        mensaje1.append(
                            {"error": "El usuario " + new_student['USERNAME'][i] + " contiene espacios en blanco, por favor verifique."})
                    else:
                        conteo = 0 + conteo

                for i in range(len(new_student)):
                    if Programas.objects.filter(programa_name=new_student['CARRERA_ID'][i]):
                        conteo = 0 + conteo
                    else:
                        conteo = 1 + conteo
                        mensaje1.append(
                            {"error": "La carrera ingresada para el usuario: " + new_student['USERNAME'][i] + " no existe"})
                
                for i in range(len(new_student)):
                    if Periodos.objects.filter(periodo=new_student['PERIODO'][i]):
                        conteo = 0 + conteo
                    else:
                        conteo = 1 + conteo
                        mensaje1.append(
                            {"error": "El periodo ingresado para el usuario: " + new_student['USERNAME'][i] + " no existe"})

                for i in range(len(new_student)):
                    if len(str(new_student['TELEFONO'][i])) == 7 or len(str(new_student['TELEFONO'][i])) == 10:
                        conteo = 0 + conteo

                    else:
                        conteo = 1 + conteo
                        mensaje1.append(
                            {"error": "El usuario " + new_student['USERNAME'][i] + " debe tener 7(fijo) o 10(celular) números, usted ingresó: " + str(new_student['TELEFONO'][i])})

                for i in range(len(new_student)):
                    ced = str(new_student['CEDULA'][i])
                    if ced[0] == 0:
                        conteo = 1 + conteo
                        mensaje1.append(
                            {"error": "El usuario " + new_student['USERNAME'][i] + " tiene un numero de documento que inicia por cero, por favor valide."})

                    else:
                        conteo = 0 + conteo

                for i in range(len(new_student)):
                    ced = str(new_student['CEDULA_ACUDIENTE'][i])
                    if ced[0] == 0:
                        conteo = 1 + conteo
                        mensaje1.append({"error": "El acudiente del usuario " +
                                        new_student['USERNAME'][i] + " tiene un numero de documento que inicia por cero, por favor valide."})

                    else:
                        conteo = 0 + conteo

                for i in range(len(new_student)):
                    if len(str(new_student['TELEFONO_ACUDIENTE'][i])) == 7 or len(str(new_student['TELEFONO_ACUDIENTE'][i])) == 10:
                        conteo = 0 + conteo
                    else:
                        conteo = 1 + conteo
                        mensaje1.append({"error": "El teléfono del acudiente del usuario " +
                                        new_student['USERNAME'][i] + " debe tener 7(fijo) o 10(celular) números, usted ingresó: " + str(new_student['TELEFONO_ACUDIENTE'][i])})

                for i in range(len(new_student)):
                    if (fecha_actual.year - new_student['NACIMIENTO'][i].year) <= 13:
                        conteo = 1 + conteo
                        mensaje1.append(
                            {"error": "El usuario " + new_student['USERNAME'][i] + " no se puede matricular, debe de tener más de 13 años"})

                    else:
                        conteo = 0 + conteo

                for i in range(len(new_student)):
                    if not CatalogsTypesDocuement.objects.filter(nombre=new_student['T_DOCUMENTO'][i]):
                        conteo = 1 + conteo
                        mensaje1.append({"error": "El usuario " + new_student['USERNAME'][i] + " tiene un tipo de documento incorrecto: " +
                                        new_student['T_DOCUMENTO'][i] + ", seleccione uno de la lista desplegable."})
                    else:
                        conteo = 0 + conteo

                for i in range(len(new_student)):
                    if not CatalogsSede.objects.filter(sede=new_student['SEDE'][i]):
                        conteo = 1 + conteo
                        mensaje1.append(
                            {"error": "Usted ingresó la sede " + new_student['SEDE'][i] + " para el usuario " + new_student['USERNAME'][i] + ", la cual no existe" + ", seleccione una sede de la lista desplegable."})

                    else:
                        conteo = 0 + conteo

                for i in range(len(new_student)):
                    if not re.search(r"^[A-Za-z0-9_!#$%&'*+\/=?`{|}~^.-]+@[A-Za-z0-9.-]+$", new_student['EMAIL'][i]):
                        conteo = 1 + conteo
                        mensaje1.append(
                            {"error": "El usuario " + new_student['USERNAME'][i] + " tiene datos que no son emails " + new_student['EMAIL'][i]})

                    else:
                        conteo = 0 + conteo

                if len(mensaje1) > 0:
                    response = JsonResponse(mensaje1, safe=False)
                    response.status_code = 400
                    return response

                # si pasa las validaciones, se comienza a procesar
                else:
                    existing_slugs = set()
                    for i in range(len(new_student)):
                        carrera = Programas.objects.get(
                            programa_name=new_student['CARRERA_ID'][i]).id
                        costo_cierre = Programas.objects.get(
                            programa_name=new_student['CARRERA_ID'][i]).costo
                        document = CatalogsTypesDocuement.objects.get(
                            nombre=new_student['T_DOCUMENTO'][i]).id
                        sede = CatalogsSede.objects.get(
                            sede=new_student['SEDE'][i]).id
                        periodo= Periodos.objects.get(
                            periodo=new_student['PERIODO'][i]).id
                        slug = Estudiante.objects.generate_unique_slug(new_student['USERNAME'][i], existing_slugs)
                        data_list.append(Estudiante(codigo=int(codigo) + i, cedula=new_student['CEDULA'][i],
                                                    nombre=new_student['NOMBRE'][i], apellidos=new_student['APELLIDOS'][i],
                                                    nacionalidad=new_student['NACIONALIDAD'][i],
                                                    telefono=new_student['TELEFONO'][i], direccion=new_student['DIRECCION'][i],
                                                    nacimiento=new_student['NACIMIENTO'][i], email=new_student['EMAIL'][i],
                                                    sexo=new_student['SEXO'][i], username=new_student['USERNAME'][i],
                                                    nombre_acudiente=new_student['NOMBRE_ACUDIENTE'][
                                                        i], apellidos_acudiente=new_student['APELLIDOS_ACUDIENTE'][i],
                                                    telefono_acudiente=new_student['TELEFONO_ACUDIENTE'][
                                                        i], cedula_acudiente=new_student['CEDULA_ACUDIENTE'][i],
                                                    carrera_id=carrera, masivo=True, periodo_matriculado_id=periodo, tDocument_id=document,
                                                    sede_id=sede, costo_cierre=costo_cierre,
                                                    slug=slug,
                                                    ))
                        User.objects.create_user(codigo=int(codigo) + i, nombres=new_student['NOMBRE'][i], apellidos=new_student['APELLIDOS'][i],
                                                 email=new_student['EMAIL'][i], username=new_student['USERNAME'][i], tipe=CatalogsTypesRol.objects.get(
                                                     rol="Estudiante"),
                                                 password=Estudiante.objects.get_secret("RANDOM"), is_superuser=False, is_active=True, is_staff=False

                                                 )

                    Estudiante.objects.bulk_create(data_list)

                    return HttpResponseRedirect(
                        reverse(
                            'student_app:list-student'
                        )
                    )
        else:
            mensaje1.append({"error":'Archivo inválido, recuerde que el archivo tiene por nombre "cargue_estudiantes.xlsx" , por favor verifique y cárguelo nuevamente'})
            response = JsonResponse(mensaje1, safe=False)
            response.status_code = 400
            return response


# Vista para listar estudiantes creados de forma masiva ok
@method_decorator(cache_control(no_cache=True, must_revalidate=True, no_store=True), name='dispatch')
class StudentCargueListview(AdminRequiredMixin,ListView):
    model = Estudiante
    template_name = 'estudiantes/masivos.html'
    context_object_name = 'student'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['programas'] = Programas.objects.all()
        return context

    def get_queryset(self):
        queryset = Estudiante.objects.filter(
                is_active=True).exclude(masivo=False)
        return queryset
    
#Vista para actualizar estudiantes creados de manera masiva, la diferencia entre esta vista y la siguiente es que en esta
    # se crean las facturas, confirmando que si se va a matricular el estudiante, en el otro solo actualizamos datos
class StudentMasiveUpdateView(AdminRequiredMixin,UpdateView):
    model = Estudiante
    template_name = 'estudiantes/update_student.html'
    form_class = StudentUpdateForm
    success_url = reverse_lazy('student_app:list-student')


    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        cod_estudiante = Estudiante.objects.get(pk=self.kwargs['pk']).codigo
        estudiante_c = self.model.objects.get(codigo=cod_estudiante)
        form = StudentUpdateForm(request.POST, instance=estudiante_c)

        if form.is_valid():
            inv_list =[]
            form.save()
            crea_user = User.objects.filter(
                codigo=cod_estudiante
            ).update(
                     email=form.cleaned_data['email'])
            act = Estudiante.objects.filter(
                pk=self.kwargs['pk']).update(masivo=False)
            
            
            programa = Programas.objects.get(programa_name=form.cleaned_data.get('carrera'))
            cuotas = int(programa.cuotas)

            if programa.tiene_grado:
                # Calcular la fecha de vencimiento para el certificado de grado
                fechafacturacertificados = datetime.now() + relativedelta(months=cuotas)
                fechafacturacertificados = fechafacturacertificados.replace(day=5)

                # Crear factura de matrícula
                Facturas.objects.create(
                    user=User.objects.get(username=form.cleaned_data.get('username')),
                    codigo=Facturas.objects.code_invoice(),
                    email=form.cleaned_data.get('email'),
                    monto=programa.matricula,
                    due_at= datetime.now(),
                    descripcion="Matrícula",
                    estado=CatalogsTypesInvoices.objects.get(estado="Pendiente"),
                )

                # Crear facturas de mensualidades
                for i in range(cuotas):
                    if i == 0:
                        fechafacturamensualidades = datetime.now()
                    else:
                        fechafacturamensualidades = datetime.now() + relativedelta(months=i)
                        fechafacturamensualidades = fechafacturamensualidades.replace(day=5)

                    inv_list.append(Facturas(
                        user=User.objects.get(username=form.cleaned_data.get('username')),
                        codigo=int(Facturas.objects.code_invoice()) + i,
                        email=form.cleaned_data.get('email'),
                        monto=programa.cuota_valor,
                        due_at=fechafacturamensualidades,
                        descripcion="Mensualidad número " + str(i + 1),
                        estado=CatalogsTypesInvoices.objects.get(estado="Pendiente"),
                    ))

                # Crear todas las facturas de mensualidades en bulk
                Facturas.objects.bulk_create(inv_list)

                # Crear factura de derechos de grado
                Facturas.objects.create(
                    user=User.objects.get(username=form.cleaned_data.get('username')),
                    codigo=Facturas.objects.code_invoice(),
                    email=form.cleaned_data.get('email'),
                    monto=programa.derechosGrado,
                    due_at=fechafacturacertificados,
                    descripcion="Derechos de Grado",
                    estado=CatalogsTypesInvoices.objects.get(estado="Pendiente"),
                )

            else:
                # Calcular la fecha de vencimiento para el certificado de grado
                fechafacturacertificados = datetime.now() + relativedelta(months=cuotas)
                fechafacturacertificados = fechafacturacertificados.replace(day=5)

                # Crear factura de inscripción
                Facturas.objects.create(
                    user=User.objects.get(username=form.cleaned_data.get('username')),
                    codigo=Facturas.objects.code_invoice(),
                    email=form.cleaned_data.get('email'),
                    monto=programa.matricula,
                    due_at= datetime.now(),
                    descripcion="Inscripción",
                    estado=CatalogsTypesInvoices.objects.get(estado="Pendiente"),
                )

                # Crear facturas de mensualidades
                for i in range(cuotas):
                    if i == 0:
                        fechafacturamensualidades = datetime.now()
                    else:
                        fechafacturamensualidades = datetime.now() + relativedelta(months=i)
                        fechafacturamensualidades = fechafacturamensualidades.replace(day=5)
                    
                    inv_list.append(Facturas(
                        user=User.objects.get(username=form.cleaned_data.get('username')),
                        codigo=int(Facturas.objects.code_invoice()) + i,
                        email=form.cleaned_data.get('email'),
                        monto=programa.cuota_valor,
                        due_at=fechafacturamensualidades,
                        descripcion="Mensualidad número " + str(i + 1),
                        estado=CatalogsTypesInvoices.objects.get(estado="Pendiente"),
                    ))

                # Crear todas las facturas de mensualidades en bulk
                Facturas.objects.bulk_create(inv_list)

                # Crear factura de certificado
                Facturas.objects.create(
                    user=User.objects.get(username=form.cleaned_data.get('username')),
                    codigo=Facturas.objects.code_invoice(),
                    email=form.cleaned_data.get('email'),
                    monto=programa.derechosGrado,
                    due_at=fechafacturacertificados,
                    descripcion="Certificado",
                    estado=CatalogsTypesInvoices.objects.get(estado="Pendiente"),
                )

            mensaje = f'{self.model.__name__} registrado correctamente'
            error = "No hay error!"
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 201
            return response
        else:
            mensaje = f'{self.model.__name__} no se ha podido registrar'
            error = form.errors
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 400
            return response

#Vista para actualizar estudiantes creados de manera normal, uno a uno   
class StudentUpdateView(AdminRequiredMixin,UpdateView):
    model = Estudiante
    template_name = 'estudiantes/update_student_normal.html'
    form_class = StudentUpdateForm
    success_url = reverse_lazy('student_app:list-student')


    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        cod_estudiante = Estudiante.objects.get(pk=self.kwargs['pk']).codigo
        estudiante_c = self.model.objects.get(codigo=cod_estudiante)
        form = StudentUpdateForm(request.POST, instance=estudiante_c)

        if form.is_valid():
            form.save()
            crea_user = User.objects.filter(
                codigo=cod_estudiante
            ).update(
                     email=form.cleaned_data['email'])

            mensaje1.append({"error": 'Registrado correctamente'})
            response = JsonResponse(mensaje1, safe=False)
            response.status_code = 201
            return response
        else:
            mensaje1 =[]
            mensaje1.append(
                {"error": form.errors}
            )
            response = JsonResponse(mensaje1, safe=False)
            response.status_code = 400
            return response
        

# Vista para generar plantilla de estudiantes, se encuentra ok


def export_users_csv(request): 

    # Variables utilizadas para traerse la información creada en modelos y catálogos
    program = Programas.objects.all()
    periods = Periodos.objects.all()
    documents = CatalogsTypesDocuement.objects.all()
    sede = CatalogsSede.objects.all()
    wb = Workbook()

    # Nombramos las pestanas que va a llevar el archivo
    ws1 = wb.create_sheet(index=0, title="Plantilla")
    ws = wb.create_sheet(index=1, title="Campos")

    # Generamos los datos que van a las celdas, y que sirven de listas desplegables
    for number in range(0, len(program)):
        ws['A{}'.format(number+1)].value = "{}".format(program[number])

    for number in range(0, len(COUNTRIES)):
        ws['B{}'.format(number+1)].value = "{}".format(COUNTRIES[number][1])

    for number in range(0, len(GENEROS)):
        ws['C{}'.format(number+1)].value = "{}".format(GENEROS[number][1])

    for number in range(0, len(documents)):
        ws['D{}'.format(number+1)].value = "{}".format(documents[number])

    for number in range(0, len(sede)):
        ws['E{}'.format(number+1)].value = "{}".format(sede[number])
    
    for number in range(0, len(periods)):
        ws['F{}'.format(number+1)].value = "{}".format(periods[number])

    # Asignamos los nombres de los encabezados a las columnas del archivo de excel
    for number in range(0, len(EST_STUDENT)):
        c1 = ws1.cell(row=1, column=number + 1)
        c1.value = EST_STUDENT[number]

    # Asignamos las listas de validación para el excel
    data_val1 = DataValidation(
        type="list", formula1='=Campos!$A$1:$A$' + str(len(program)))
    data_val2 = DataValidation(
        type="list", formula1='=Campos!$B$1:$B$' + str(len(COUNTRIES)))
    data_val3 = DataValidation(
        type="list", formula1='=Campos!$C$1:$C$' + str(len(GENEROS)))
    data_val4 = DataValidation(
        type="list", formula1='=Campos!$D$1:$D$' + str(len(documents)))
    data_val5 = DataValidation(
        type="list", formula1='=Campos!$E$1:$E$' + str(len(sede)))
    data_val6 = DataValidation(
        type="list", formula1='=Campos!$F$1:$F$' + str(len(periods)))

    # Asignamos las listas desplegables al la hoja principal
    ws1.add_data_validation(data_val1)
    ws1.add_data_validation(data_val2)
    ws1.add_data_validation(data_val3)
    ws1.add_data_validation(data_val4)
    ws1.add_data_validation(data_val5)
    ws1.add_data_validation(data_val6)

    data_val1.add(ws1["P2"])
    data_val2.add(ws1["E2"])
    data_val3.add(ws1["J2"])
    data_val4.add(ws1["A2"])
    data_val5.add(ws1["Q2"])
    data_val6.add(ws1["R2"])

    content = save_virtual_workbook(wb)
    response = HttpResponse(content)
    response['Content-Disposition'] = 'attachment; filename=cargue_estudiantes.xlsx'
    response['Content-Type'] = 'application/x-xlsx'
    return response

# Vista se encuentra ok, elimina estudiantes.
class StudentDeleteView(AdminRequiredMixin,DeleteView):
    template_name = 'estudiantes/delete_student.html'
    model = Estudiante
    success_url = reverse_lazy('student_app:list-student')

    def post(self, request, pk, *args, **kwargs):
        consulta1 = Estudiante.objects.get(codigo=pk).delete()
        consulta2 = User.objects.get(codigo=pk).delete()

        return HttpResponseRedirect(self.success_url)
    
# Vista que muestra el detalle del estudiante
class StudentDetailView(DetailView):
    template_name = 'estudiantes/detail_student.html'
    model = Estudiante

#Vista que muestra las materias que se pueden asignar al estudiante
class StudentAssignListview(AdminRequiredMixin,ListView):
    model = Estudiante
    template_name = 'estudiantes/AssignCreate.html'
    success_url = reverse_lazy('student_app:list-student')


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        estudiante = Estudiante.objects.get(pk=self.kwargs['pk'])
        asignaturas = Materias.objects.filter(materia__programa=estudiante.carrera, is_active=True)
        datos = []
        for i in asignaturas:
            total = Banner.objects.filter( materia_id= i.pk ).values("student_id").distinct().count()
            
            if Banner.objects.filter(student_id=estudiante.pk, materia_id= i.pk ).exists():
                data_json = {"pk": i.id, "codigo": i.materia.codigo, "nombres": i.materia.nombre_materia,
                                "docente":i.docente, "jornada": i.jornada, "periodo": i.periodo, 
                                "total":total, "estado": "Asignada"}
                datos.append(data_json)
            else:
                data_json = {"pk": i.id, "codigo": i.materia.codigo, "nombres": i.materia.nombre_materia,
                                "docente":i.docente, "jornada": i.jornada, "periodo": i.periodo, 
                                "total":total, "estado": "Sin asignar"}
                datos.append(data_json)
        
        context['asignaturas'] = datos
        context['estudiante'] = estudiante
        return context


#Vista que muestra las asignaturas que ya tiene asignado el estudiante
@method_decorator(cache_control(no_cache=True, must_revalidate=True, no_store=True), name='dispatch')
class StudentNotesListview(ListView):
    model = Banner
    template_name = 'estudiantes/ListNotesView.html'
    context_object_name = 'student'
    success_url = reverse_lazy('student_app:list-student')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        estudiante = Estudiante.objects.get(slug=self.kwargs['slug'])
        context['estudiante'] = estudiante

        datos = Banner.objects.filter(student__slug=self.kwargs['slug'], materia__is_active=False).values("materia_id", 
                                                                           "materia__an_creacion", 
                                                                           "materia__materia__codigo").distinct().order_by(
            'materia__materia__codigo','-materia__an_creacion'
        )
        dataLista=[]
        mean=[]
        pasadas=[]
        for i in datos:
            dataLista.append({"materia_id":i['materia_id'],"fecha": i['materia__an_creacion'], "codigo":i['materia__materia__codigo']})

        df = pd.DataFrame(dataLista)
        df["duplicado"]=df.duplicated("codigo")

        values = [False]
        df = df[df.duplicado.isin(values)]
        if len(df)>0:
            df = df['materia_id']

            for i in df:
                promedio = round (Banner.objects.filter(student__slug=self.kwargs['slug'], materia_id=i ).aggregate(Avg('calificacion'))['calificacion__avg'],2)
                mean.append( promedio )
                if promedio >= 3.0:
                    pasadas.append(1)
                else:
                    pasadas.append(0)
            mean = round(sum(mean) / len(mean),2)
            final = int(round((sum(pasadas)/int(estudiante.carrera.aceptado))*100,0))
            context['promedio'] = mean
            context['pasadas'] = final
            return context
        else:
            context['promedio'] = 0.00
            context['pasadas'] = 0
        #promedio, cumplimiento
            return context

    def get_queryset(self):

        data = []
    
        # Obtener los datos sin usar distinct(), pero usando values() para reducir la carga de datos
        estudiante = Banner.objects.filter(student__slug=self.kwargs['slug']).values(
            'student__slug', 'materia_id', 'materia__materia__codigo', 
            'materia__materia__nombre_materia', 'materia__is_active', 
            'materia__periodo'
        ).order_by('student__slug', 'materia_id')

        # Crear un diccionario para asegurarnos de que no haya duplicados por 'materia_id'
        materias_vistas = set()

        for i in estudiante:
            # Verificar si ya hemos procesado la materia
            if i['materia_id'] not in materias_vistas:
                # Agregar la materia al set para evitar duplicados
                materias_vistas.add(i['materia_id'])
                
                # Calcular el promedio de calificaciones para esta materia
                promedio = round(
                    Banner.objects.filter(student__slug=self.kwargs['slug'], materia_id=i['materia_id']).aggregate(Avg('calificacion'))['calificacion__avg'], 
                    2
                )
                
                # Crear el diccionario con los datos de la materia
                data_json = {
                    'slug': i['student__slug'], 
                    'pkmateria': i['materia_id'], 
                    "codigo": i['materia__materia__codigo'], 
                    "nombre": i['materia__materia__nombre_materia'],
                    "estado": i['materia__is_active'], 
                    "promedio": promedio, 
                    'periodo':  Periodos.objects.get(pk=i['materia__periodo']).periodo,
                    "tiene": "si" if promedio != 0.00 else "no"
                }
                
                # Agregar los datos al arreglo
                data.append(data_json)

        return data
    

#Vista que muestra el detalle de notas de la asignatura seleccionada de el estudiante
@method_decorator(cache_control(no_cache=True, must_revalidate=True, no_store=True), name='dispatch')
class StudentNotesDetailListview(ListView):
    model = Banner
    template_name = 'estudiantes/ListNotesDetailView.html'
    context_object_name = 'student'
    success_url = reverse_lazy('student_app:list-student')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        estudiante_data = Estudiante.objects.get(slug=self.kwargs['slug'])
        context['estudiante'] = estudiante_data
        data = [] 
        estudiante = Materias.objects.filter(pk=self.kwargs['pk1'] )

        for i in estudiante:
            promedio = round (Banner.objects.filter(student__slug=self.kwargs['slug'], materia_id=self.kwargs['pk1'] ).aggregate(Avg('calificacion'))['calificacion__avg'],2)
            
            data_json = {
                'pk': i.pk, "codigo": i.materia.codigo, "nombre": i.materia.nombre_materia,
                "estado": i.is_active, "promedio": promedio
            }
            data.append(data_json)
        context['informacion'] = data
        return context
    
    def get_queryset(self):
        data = []
        estudiante = Banner.objects.filter(materia_id=self.kwargs['pk1'], student__slug=self.kwargs['slug']  )

        for i in estudiante:
            datos = {"estudiante": i.student.nombre + " " + i.student.apellidos, "cod_tarea": i.cod_tarea, "tarea": i.tarea.tipo, "calificacion": i.calificacion }
            data.append(datos)
        estudiante = pd.DataFrame(data, columns=['estudiante', 'tarea', "calificacion", "cod_tarea"])
        estudiante["cod_tareaG"] = estudiante["cod_tarea"] + "-" + estudiante["tarea"]
        estudiante = estudiante.pivot(index='estudiante', columns='cod_tareaG', values='calificacion')

        return estudiante

#Vista que elimina a los estudiantes de los salones
class StudentNotesDeleteview(DeleteView):
    template_name = 'estudiantes/StudentNotesDeleteview.html'
    model = Banner

    def get_object(self, queryset=None):
        if queryset is None:
            queryset = self.get_queryset()

        materias = self.kwargs['pk1']
        estudiante = self.kwargs['slug']
        context = {'materias':materias, 'estudiante':estudiante}
        return context
    

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["estudiante"] = Estudiante.objects.get(slug=self.kwargs['slug'] )
        context["materias"] = Materias.objects.get(pk=self.kwargs['pk1'] )
        return context
    
    def post(self, request, *args, **kwargs):
        estudiante = Banner.objects.filter(materia_id=self.kwargs['pk1'], 
                                           student_id=Estudiante.objects.get(slug=self.kwargs['slug'] ).pk  ).delete()

        return HttpResponseRedirect(reverse_lazy('student_app:list-student'))
    

#---------Desde aqui comenzamos con todo lo relacionado al grado ------------------------------------


# Con esta vista mostramos los estudiantes aptos para graduarse, se encuentra ok
@method_decorator(cache_control(no_cache=True, must_revalidate=True, no_store=True), name='dispatch')
class StudentListFinalily(ListView):
    model = Banner
    second_model = Programas
    template_name = 'estudiantes/list_aptos.html'
    context_object_name = 'finaly'

    def get_queryset(self):
        estudiante_list = []
        #Ordenamos las asignaturas que tienen un promedio mayor a 2.99

        aceptado = Banner.objects.filter(
            materia__is_active=False  # Solo materias cerradas
        ).values('student','materia').annotate(
            promedio_calificacion=Avg('calificacion') # Contamos la cantidad de materias por estudiante
        ).filter(
            promedio_calificacion__gt=2.99  # Solo aquellos estudiantes con promedio mayor a 2.99
        )

        estudiantes = [item['student'] for item in aceptado]

        # Contar cuántas veces aparece cada estudiante usando una list comprehension y dict comprehension
        aceptado = [{'student__id': estudiante, 'materias_aprobadas': estudiantes.count(estudiante)} for estudiante in set(estudiantes)]

        
            
        for i in list(aceptado):
            st = Estudiante.objects.get(id=i['student__id'])
            if Graduated.objects.filter(student_id=st.pk).count() == 0:
                carrera = Programas.objects.get(id=st.carrera_id)
                pagadas = Facturas.objects.filter(
                    user_id=st.codigo, estado_id=3).count()
                rr = {'pk': st.pk, 'slug':st.slug, 'codigo': st.codigo, 'estudiante': st.apellidos + ' ' + st.nombre, 'programa': carrera.programa_name,
                      'indicador': int((int(i['materias_aprobadas'])/int(carrera.aceptado))*100), 'financiero': int((int(pagadas)/(2 + int(carrera.cuotas)))*100)}

                estudiante_list.append(rr)

        #Ordenamos los resultados antes de enviarlos al html
        def get_age(indica):
            return indica.get('indicador')

        estudiante_list.sort(key=get_age, reverse=True)
        return estudiante_list


#Con esta vista graduamos a los estudiantes, se puede hacer de uno en uno o por multiples estudiantes **
class StudentGraduateView(AdminRequiredMixin, CreateView):
    model = Graduated
    form_class = GraduateRegisterForm
    template_name = 'estudiantes/list_aptos.html'
    success_url = reverse_lazy('student_app:list-student')

    def get(self, request, *args, **kwargs):
        form = GraduateRegisterForm()
        context = {"form": form}
        return render(request, r"estudiantes/studentGraduated.html", context)

    def post(self, request, *args, **kwargs):
        todos = []
        date = datetime.now()
        studen_graduate = str(self.request.POST.get('concat'))
        studen_folio = str(self.request.POST.get('folio'))
        studen_libro = str(self.request.POST.get('libro'))
        studen_graduate = studen_graduate.split(sep=",")

        for i in studen_graduate:
            consulta = Estudiante.objects.get(pk=i)
            individual = Graduated(student_id=consulta.pk, libro=studen_libro, folio=studen_folio)
            todos.append(individual)
            Estudiante.objects.filter(id=int(i)).update(is_active=False, is_graduado=True)
        Graduated.objects.bulk_create(todos)
        return HttpResponseRedirect(
            self.request.META.get("HTTP_REFERER")

        )
    

#Listar graduados
@method_decorator(cache_control(no_cache=True, must_revalidate=True, no_store=True), name='dispatch')
class GraduatedListView(ListView):
    model = Graduated
    template_name = 'estudiantes/GraduadoList.html'
    context_object_name = 'graduados'

    def get_queryset(self):
        graduatedList=[]
        dataGraduated = Graduated.objects.all()
        for i in dataGraduated:
            TotalpromedioSum = Banner.objects.filter(student_id=i.student_id).aggregate(Avg('calificacion'))
            Totalpromedio = round( TotalpromedioSum['calificacion__avg'],2)
            data = {'estudiante_id': i.student_id, 'codigo': i.student.codigo, 
                    'estudiante': i.student.apellidos + ' ' + i.student.nombre, 
                    'programa': i.student.carrera.programa_name,
                      'promedio': Totalpromedio, 'pk': i.pk}

            graduatedList.append(data)
        return graduatedList



#La siguiente vista elimina estudiantes de la zona de graduados y los devuelve a ser estudiantes activos


class StudentGraduatedDeleteview(DeleteView):
    template_name = 'estudiantes/StudentGraduatedDeleteview.html'
    model = Graduated

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["estudiante"] = Estudiante.objects.get(pk=Graduated.objects.get(pk=self.kwargs['pk']).student_id )
        return context

    def post(self, request, *args, **kwargs):
        estudiante2 = Estudiante.objects.get(pk=Graduated.objects.get(pk=self.kwargs['pk']).student_id)
         # Actualizar los campos
        estudiante2.is_active = True
        estudiante2.is_graduado = False

        # Guardar los cambios
        estudiante2.save()
        estudiante = Graduated.objects.get(pk=self.kwargs['pk'] ).delete()

        return HttpResponseRedirect(reverse_lazy('student_app:list-graduated'))