
from decimal import *
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook
import warnings
import pandas as pd


from django.http import JsonResponse, HttpResponse
from django.views.generic import (TemplateView,
    FormView, CreateView, DeleteView, UpdateView,
    DetailView, ListView, View)
from django.urls import reverse_lazy, reverse
from django.shortcuts import render, HttpResponseRedirect
from django.db.models import F, Sum, Avg, Count

from .models import *
from .forms import *





# Vista ok
class ConfigTemplateView(TemplateView):
    template_name = "configuraciones/setings.html"


#------- periodos ---------
class PeriodoCreateView(CreateView):
    model = Periodos
    form_class = PeriodoForm
    template_name = 'periodos/register_periodo.html'
    success_url = reverse_lazy('settings_app:list-periodo')

    def post(self, request, *args, **kwargs):
        form = PeriodoForm(self.request.POST)
        
        if form.is_valid():
            periodo = str(form.cleaned_data.get('periodo')),
            create_peri = Periodos.objects.create_periodo(
            periodo[0],
            )
            mensaje = f'{self.model.__name__} creado correctamente'
            error = "No hay error!"
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 201
            return response
        else:
            mensaje1 =[]
            mensaje1.append(
                {"error": form.errors}
            )
            response = JsonResponse(mensaje1, safe=False)
            response.status_code = 400
            return response

class PeriodoListView(ListView):
    model = Periodos
    template_name = 'periodos/list_periodos.html'
    context_object_name = 'periodos'

    def get_queryset(self):

        data_periodos = Periodos.objects.all()
        data = []
        for i in data_periodos:
            if Materias.objects.filter(periodo_id=i.id):
                data_json = {"id": i.id, "periodo": i.periodo, "estado": True}
                data.append(data_json)
            else:
                data_json = {"id": i.id, "periodo": i.periodo, "estado": False}
                data.append(data_json)

        return data


class PeriodoDeleteView(DeleteView):
    template_name = 'periodos/detele_periodo.html'
    model = Periodos
    success_url = reverse_lazy('settings_app:list-periodo')




#------- programas ---------


#Vista de creacion de programas ok
class ProgramaCreateView(CreateView):
    model = Programas
    form_class = ProgramaForm
    template_name = 'programas/register_program.html'
    success_url = reverse_lazy('settings_app:list-program')

    def post(self, request, *args, **kwargs):

        form = ProgramaForm(self.request.POST)
        
        if form.is_valid():

            crear_programa = Programas.objects.create(
                tipe=form.cleaned_data['tipe'], 
                cod_prog=Programas.objects.code_programas(),
                matricula=form.cleaned_data['matricula'],
                cuota_valor=form.cleaned_data['cuota_valor'],
                cuotas=form.cleaned_data['cuotas'],
                costo=form.cleaned_data['costo'],
                derechosGrado=form.cleaned_data['derechosGrado'],
                programa_name=form.cleaned_data['programa_name'],
                aceptado=form.cleaned_data['aceptado']

            )
            mensaje = f'{self.model.__name__} creado correctamente'
            error = "No hay error!"
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 201
            return response
        else:
            mensaje1 =[]
            mensaje1.append(
                {"error": form.errors}
            )
            response = JsonResponse(mensaje1, safe=False)
            response.status_code = 400
            return response


#Vista de detalle de programas ok
class ProgramaDetailView(DetailView):
    template_name = 'programas/detail_program.html'
    model = Programas


    def get_context_data(self, **kwargs):
        context = super( ProgramaDetailView, self).get_context_data(**kwargs)
        datos = Programas.objects.get(
            pk=self.kwargs['pk'])
        datos_final = {"cod_prog": datos.cod_prog, 'programa_name': datos.programa_name,
                'matricula': f'$ {datos.matricula:,.2f}', 'costo': f'$ {datos.costo:,.2f}', 'cuota_valor': f'$ {datos.cuota_valor:,.2f}',
                'Dgrado': f'$ {datos.derechosGrado:,.2f}', "cuotas": datos.cuotas
            }

        context['datos'] = datos_final
        return context



#Vista de eliminacion de programas ok
class ProgramaDeleteView(DeleteView):
    template_name = 'programas/delete_program.html'
    model = Programas
    success_url = reverse_lazy('settings_app:list-program')

#Vista de actualizacion de programas ok
class ProgramaUpdateView(UpdateView):
    template_name = 'programas/update_program.html'
    model = Programas
    form_class = ProgramaForm
    success_url = reverse_lazy('settings_app:list-program')

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        user_c = self.model.objects.get(pk=self.kwargs['pk'])
        form = ProgramaForm(request.POST, instance=user_c)
        
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} Actualizado correctamente'
            error = "No hay error!"
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 201
            return response
        else:
            mensaje1 =[]
            mensaje1.append(
                {"error": form.errors}
            )
            response = JsonResponse(mensaje1, safe=False)
            response.status_code = 400
            return response

#Vista de listado de programas ok
class Programalistview(ListView):
    model = Programas
    template_name = 'programas/list_programas.html'
    context_object_name = 'programas'

    def get_queryset(self):
        kword = self.request.GET.get('kword', '')
        data_program = []
        if kword:
            data_prin = Programas.objects.filtrar_buscador(kword)
            for i in data_prin:
                if Materias.objects.filter(materia__programa_id=i.id).exists():
                    data_json = {"pk": i.id, "codigo": i.cod_prog,
                                 "programa": i.programa_name, "estado": True, "is_active": i.is_active}
                    data_program.append(data_json)
                else:
                    data_json = {"pk": i.id, "codigo": i.cod_prog,
                                 "programa": i.programa_name, "estado": False, "is_active": i.is_active}
                    data_program.append(data_json)
            queryset = data_program

        else:

            data_prin = Programas.objects.all()
            for i in data_prin:
                if Materias.objects.filter(materia__programa_id=i.id).exists():
                    data_json = {"pk": i.id, "codigo": i.cod_prog,
                                 "programa": i.programa_name, "estado": True, "is_active": i.is_active}
                    data_program.append(data_json)
                else:
                    data_json = {"pk": i.id, "codigo": i.cod_prog,
                                 "programa": i.programa_name, "estado": False, "is_active": i.is_active}
                    data_program.append(data_json)
            queryset = data_program

        return queryset





#------- Inventario asignaturas ---------
    

class Inventariolistview(ListView):
    model = Inventario
    template_name = 'inventario/inventarioList.html'
    context_object_name = 'inventario'

    def get_queryset(self):

        data_Inventari = Inventario.objects.all()
        data = []
        for i in data_Inventari:
            if Materias.objects.filter(materia_id=i.id):
                data_json = {"pk": i.id, "codigo": i.codigo, "estado": True, 
                             "programa_name": i.programa.programa_name,
                            "nombre_materia": i.nombre_materia }
                data.append(data_json)
            else:
                data_json = {"pk": i.id, "codigo": i.codigo, "estado": False, 
                             "programa_name": i.programa.programa_name,
                            "nombre_materia": i.nombre_materia }
                data.append(data_json)

        return data


#Vista de creacion de programas ok
class InventarioCreateView(CreateView):
    model = Inventario
    form_class = InventarioRegisterForm
    template_name = 'inventario/inventarioRegister.html'
    success_url = reverse_lazy('settings_app:list-inventario')

    def post(self, request, *args, **kwargs):

        form = InventarioRegisterForm(self.request.POST)
        
        if form.is_valid():
            crear_asignaturas = Inventario.objects.create(
                codigo = Inventario.objects.code_asignaturas(),
                nombre_materia=form.cleaned_data['nombre_materia'],
                programa=form.cleaned_data['programa']

            )
            mensaje = f'{self.model.__name__} creado correctamente'
            error = "No hay error!"
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 201
            return response
        else:
            mensaje1 =[]
            mensaje1.append(
                {"error": form.errors}
            )
            response = JsonResponse(mensaje1, safe=False)
            response.status_code = 400
            return response


#Vista de actualizacion de programas ok
class InventarioUpdateView(UpdateView):
    model = Inventario
    template_name = 'inventario/inventarioUpdate.html'
    form_class = InventarioRegisterForm
    success_url = reverse_lazy('settings_app:list-inventario')

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        user_c = self.model.objects.get(pk=self.kwargs['pk'])
        form = InventarioRegisterForm(request.POST, instance=user_c)
        
        if form.is_valid():
            form.save()
            mensaje = f'{self.model.__name__} Actualizado correctamente'
            error = "No hay error!"
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 201
            return response
        else:
            mensaje1 =[]
            mensaje1.append(
                {"error": form.errors}
            )
            response = JsonResponse(mensaje1, safe=False)
            response.status_code = 400
            return response
        

class InventarioDeleteView(DeleteView):
    template_name = 'inventario/inventarioDelete.html'
    model = Inventario
    success_url = reverse_lazy('settings_app:list-inventario')

#class InventarioMasiveCreateView(CreateView):

class InventarioMasiveView(View):

    def get(self, request, *args, **kwargs):
        form = InventarioMasiveForm()
        context = {"form": form}
        return render(request, r"inventario\inventarioMasive.html", context)
    
    def post(self, request, *args, **kwargs):

        listaEncabezados = ["Nombre_asignatura", "Programa"]
        mensaje =[]
        data_list =[]
        conteo = 0
        # Leyendo el archivo que recibimos desde el form
        Data = self.request.FILES['carga']

        # Validacion 1
        if Data.name == 'cargue_asignaturas.xlsx':
            warnings.filterwarnings(
                'ignore', category=UserWarning, module='openpyxl')
            newData = pd.read_excel(
                Data, sheet_name="Plantilla", engine='openpyxl')
            


            # Validacion 1.1 valores en blanco y encabezados ok
            # Validamos que no existan valores nulos y en blancos en el archivo.
            missValues = newData.isnull().sum()
            missValues =  missValues[ missValues != 0]

            # Generamos código para la asignatura
            codigo = Inventario.objects.code_asignaturas()

            # comparando los encabezados para ver si está correcta la estructura
            for i, a in zip(list(newData), listaEncabezados):
                if i != a:
                    break

            # Desplegamos errores en caso de existir
            if missValues.shape[0] or i != a:
                if missValues.shape[0]:
                    if missValues.shape[0] == 1:
                        mensaje.append({"error": 'El archivo tiene ' + str(
                            missValues.shape[0]) + ' columna sin datos, por favor, revise'})
                        response = JsonResponse(mensaje, safe=False)
                        response.status_code = 400
                        return response
                    else:
                        mensaje.append({"error": 'El archivo tiene ' + str(
                            missValues.shape[0]) + ' columnas sin datos, por favor, revise'})
                        response = JsonResponse(mensaje, safe=False)
                        response.status_code = 400
                        return response

                else:
                    mensaje.append({"error": 'La columna con encabezado ' + i +
                                    ' en su archivo, no es válido, verifique el archivo y vuelva a cargarlo'})
                    response = JsonResponse(mensaje, safe=False)
                    response.status_code = 400
                    return response
                    
            # Validacion 1.1 mas de 20 asignaturas y archivos en blanco
            elif len(newData) > 20 or len(newData) == 0:
                if len(newData) > 20:
                    mensaje.append(
                        {"error": 'Recuerde que no puede cargar más de 20 asignaturas a la vez, en este archivo encontramos: ' + str(len(newData))})
                    response = JsonResponse(mensaje, safe=False)
                    response.status_code = 400
                    return response
                else:
                    mensaje.append(
                        {"error": 'No encontramos asignaturas para cargar'})
                    response = JsonResponse(mensaje, safe=False)
                    response.status_code = 400
                    return response
                
            else:
                newData['DUPLICADO'] = newData.duplicated()

                for i in range(len(newData)):
                    if newData['DUPLICADO'][i] == True:
                        conteo = 1 + conteo
                        mensaje.append({"error": "Encontramos filas duplicadas para la asignatura: " +
                                        newData['Nombre_asignatura'][i] + " verifique la información"})
                    else:
                        conteo = 0 + conteo
                
                for i in range(len(newData)):
                    if Programas.objects.filter(programa_name=newData['Programa'][i]):
                        conteo = 0 + conteo
                    else:
                        conteo = 1 + conteo
                        mensaje.append(
                            {"error": "El programa " + newData['Programa'][i] + " no existe, verifique la información que intenta cargar."})

                for i in range(len(newData)):
                    if not Inventario.objects.filter(programa__programa_name=newData['Programa'][i], nombre_materia= newData['Nombre_asignatura'][i] ):
                        conteo = 0 + conteo
                    else:
                        conteo = 1 + conteo
                        mensaje.append(
                            {"error": "La asignatura " + newData['Nombre_asignatura'][i] + " del programa " +newData['Programa'][i]+  "  ya se encuentra creada."})
                       
                if len(mensaje) > 0:
                    response = JsonResponse(mensaje, safe=False)
                    response.status_code = 400
                    return response

                # si pasa las validaciones, se comienza a procesar
                else:
                    for i in range(len(newData)):
                        carrera = Programas.objects.get(
                            programa_name=newData['Programa'][i]).id
                        
                        data_list.append(Inventario(
                            codigo=int(codigo) + i,
                            nombre_materia = newData['Nombre_asignatura'][i],
                            programa_id= carrera
                        ))
                    Inventario.objects.bulk_create(data_list)
                    
                    return HttpResponseRedirect(
                        reverse(
                            'settings_app:list-inventario'
                        )
                    )
        
        else:
            mensaje =[]
            mensaje.append(
                {"error": "Archivo inválido, recuerde que el archivo tiene por nombre: cargue_asignaturas.xlsx, por favor verifique y cárguelo nuevamente"}
            )
            response = JsonResponse(mensaje, safe=False)
            response.status_code = 400
            return response


# Vista para generar plantilla de creacion de asignaturas, se encuentra ok

def InventarioMasiveExport(request):

    # Variables utilizadas para traerse la información creada en modelos y catálogos
    program = Programas.objects.all()
    wb = Workbook()

    # Nombramos las pestanas que va a llevar el archivo
    ws1 = wb.create_sheet(index=0, title="Plantilla")
    ws = wb.create_sheet(index=1, title="Campos")

    # Generamos los datos que van a las celdas, y que sirven de listas desplegables
    for number in range(0, len(program)):
        ws['A{}'.format(number+1)].value = "{}".format(program[number])

    # Asignamos los nombres de los encabezados a las columnas del archivo de excel
    c1 = ws1.cell(row=1, column= 1)
    c1.value = "Nombre_asignatura"
    c2 = ws1.cell(row=1, column= 2)
    c2.value = "Programa"

    # Asignamos las listas de validación para el excel
    data_val1 = DataValidation(
        type="list", formula1='=Campos!$A$1:$A$' + str(len(program)))
    
    # Asignamos las listas desplegables al la hoja principal
    ws1.add_data_validation(data_val1)
    data_val1.add(ws1["B2"])
    

    content = save_virtual_workbook(wb)
    response = HttpResponse(content)
    response['Content-Disposition'] = 'attachment; filename=cargue_asignaturas.xlsx'
    response['Content-Type'] = 'application/x-xlsx'
    return response





#------- Vistas de asignaturas ---------


#Vista de creacion de asignaturas ok
class MateriasCreateView(CreateView):
    model = Materias
    form_class = MateriasForm
    template_name = 'materias/materiasRegister.html'
    success_url = reverse_lazy('settings_app:list-inventario')


    def get_context_data(self, **kwargs):
        context = super(MateriasCreateView, self).get_context_data(**kwargs)
        context["asignatura"] = Inventario.objects.get(pk=self.kwargs['pk'])
        context["pk"] = self.kwargs['pk']
        return context
    

    def post(self, request, *args, **kwargs):

        form = MateriasForm(self.request.POST)
        if form.is_valid():
            crear_asignaturas = Materias.objects.create(
                materia = Inventario.objects.get(pk=self.kwargs['pk']), 
                sede=form.cleaned_data['sede'],
                periodo=form.cleaned_data['periodo'],
                docente=form.cleaned_data['docente'],
                jornada=form.cleaned_data['jornada'],
                pre_cierre=form.cleaned_data['pre_cierre'],
                cierre=form.cleaned_data['cierre'],

            )
            mensaje = f'{self.model.__name__} creado correctamente'
            error = "No hay error!"
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 201
            return response
        else:
            mensaje1 =[]
            mensaje1.append(
                {"error": form.errors}
            )
            response = JsonResponse(mensaje1, safe=False)
            response.status_code = 400
            return response

#Vista que sirve para listar materias
class Materialistview(ListView):
    model = Materias
    template_name = 'materias/materiasList.html'
    context_object_name = 'materias'

    def get_context_data(self, **kwargs):
        context = super(Materialistview, self).get_context_data(**kwargs)
        context["asignatura"] = Inventario.objects.get(pk=self.kwargs['pk'])
        context["pk"] = self.kwargs['pk']
        return context

    def get_queryset(self):

        Data = Materias.objects.filter(materia_id = self.kwargs['pk'])
        datosBase = []

        for i in Data:
            total=Banner.objects.filter(materia_id = i.pk).count()
            datos = {"pk": i.pk, "sede": i.sede, "docente": i.docente, "periodo": i.periodo,
                     "jornada": i.jornada, "cierre":i.cierre, "is_active": i.is_active, "total": total }
            datosBase.append(datos)
        return datosBase
    


#------- Vistas de banner ---------
    

# Sacar las actividades distintas creadas en la asignatura filtrada
# crear un query con los estudiantes distintos de la asignatura filtrada
#cruzamos esto con las notas
    
#Vista que sirve para listar notas del salon completo
class Bannerlistview(ListView):
    model = Banner
    template_name = 'banner/bannerList.html'
    context_object_name = 'banner'


    def get_context_data(self, **kwargs):
        context = super(Bannerlistview, self).get_context_data(**kwargs)
        context["informacion"] = Materias.objects.get(pk=self.kwargs['pk'])
        
        context["general"]  = round (Banner.objects.filter(materia_id=self.kwargs['pk'] ).aggregate(Avg('calificacion'))['calificacion__avg'],2)
        return context
    

    def get_queryset(self):
        data = []
        estudiante = Banner.objects.filter(materia_id=self.kwargs['pk'])

        for i in estudiante:
            datos = {"Código": i.student.codigo, "Estudiante": i.student.nombre + " " + i.student.apellidos, 
                     "tarea": i.tarea.tipo, "calificacion": i.calificacion, "cod_tarea":i.cod_tarea
                     }
            data.append(datos)
        estudiante = pd.DataFrame(data, columns=["Código",'Estudiante', 'tarea', "calificacion","cod_tarea"])
        notas = estudiante.drop(['Estudiante','tarea',"cod_tarea"], axis=1)
        notas= notas.groupby(by='Código').agg(['mean'])
        notas.set_axis(["Promedio"], axis="columns", inplace=True)
        estudiante["Cod-Tarea"]=estudiante["cod_tarea"]+ "-" + estudiante["tarea"]
        estudiante = estudiante.pivot(index=["Código",'Estudiante'], columns=['Cod-Tarea'], values='calificacion'
                                      )
        estudiante =estudiante.reset_index()
        
        estudiante=pd.merge(estudiante, notas, on='Código')
        return estudiante

#Vista que matricula estudiantes en las asignaturas
class BannerCreateView(View):

    def post(self, request, *args, **kwargs):

        asignaturas = str(self.request.POST.get('asignaturas'))
        estudiante = str(self.request.POST.get('estudiante'))
        banner_create = asignaturas.split(sep=",")
        codigo = int(Banner.objects.code_task())
        todos = []

        for i in banner_create:
            tareas = Banner.objects.filter(materia_id= i).values("tarea_id", "cod_tarea", "observacion").distinct()
            if tareas:
                tamano = len(tareas)
                for j in range(tamano):
                    consulta = Estudiante.objects.get(pk=estudiante)
                    individual = Banner(student_id=consulta.pk, tarea_id= int(tareas[j]["tarea_id"]), 
                                        cod_tarea= tareas[j]["cod_tarea"], 
                                        observacion= tareas[j]["observacion"], 
                                        materia_id= i, calificacion= 0.0 )
                    todos.append(individual)
            else:
                
                consulta = Estudiante.objects.get(pk=estudiante)
                individual = Banner(student_id=consulta.pk, tarea_id= 6, cod_tarea= codigo, materia_id= i, calificacion= 0.0 )
                codigo += 1
                todos.append(individual)
        Banner.objects.bulk_create(todos)

        return HttpResponseRedirect(reverse_lazy('student_app:list-student'))

#Vista que crea las tareas y actividades de los estudiantes
class BannerCreateTaskView(CreateView):
    model = Banner
    form_class = BannerTaksForm
    template_name = 'banner/bannerRegisterTask.html'
    success_url = reverse_lazy('student_app:list-student')

    

    def get_context_data(self, **kwargs):
        context = super(BannerCreateTaskView, self).get_context_data(**kwargs)
        context["pk"] = self.kwargs['pk']
        return context
    

    def post(self, request, *args, **kwargs):

        form = BannerTaksForm(self.request.POST)
        if form.is_valid():

            todos = []
            todosCreate = []
            if Banner.objects.filter( materia_id= self.kwargs['pk'], tarea_id= 6).exists():
               
               estudiantes = Banner.objects.filter( materia_id= self.kwargs['pk'] ).values("student_id").distinct()
               for i in range(len(estudiantes)):
                    todos.append(estudiantes[i]["student_id"])
               todos = set(todos)
               codigo = Banner.objects.code_task()
               for i in todos:
                    
                    individual = Banner(student_id=i, 
                                        tarea_id= CatalogsTypesActivities.objects.get(tipo = form.cleaned_data['tarea']).id,
                                        materia_id= self.kwargs['pk'], calificacion= 0.0, 
                                        cod_tarea = codigo, observacion = form.cleaned_data['observacion'] )
                    
                    todosCreate.append(individual)
               Banner.objects.bulk_create(todosCreate)
               Banner.objects.filter( materia_id= self.kwargs['pk'], tarea_id= 6).delete()
               mensaje = f'{self.model.__name__} registrado correctamente'
               error = "No hay error!"
               response = JsonResponse({"mensaje": mensaje, "error": error})
               response.status_code = 201
               return response

            else:
                estudiantes = Banner.objects.filter( materia_id= self.kwargs['pk'] ).values("student_id").distinct()
                for i in range(len(estudiantes)):
                    todos.append(estudiantes[i]["student_id"])
                todos = set(todos)
                codigo = Banner.objects.code_task()
                for i in todos:
                    
                    individual = Banner(student_id=i, 
                                        tarea_id= CatalogsTypesActivities.objects.get(tipo = form.cleaned_data['tarea']).id,
                                        materia_id= self.kwargs['pk'], calificacion= 0.0, 
                                        cod_tarea = codigo, observacion = form.cleaned_data['observacion'] )
                    
                    todosCreate.append(individual)
                Banner.objects.bulk_create(todosCreate)
                mensaje = f'{self.model.__name__} registrado correctamente'
                error = "No hay error!"
                response = JsonResponse({"mensaje": mensaje, "error": error})
                response.status_code = 201
                return response
        else:
            mensaje =[]
            mensaje.append(
                {"error": form.errors}
            )
            response = JsonResponse(mensaje, safe=False)
            response.status_code = 400
            return response

class ListBannerTaskDetailView(ListView):
    model = Banner
    template_name = 'banner/listBannerTaskDetail.html'
    context_object_name = 'banner'

    def get_context_data(self, **kwargs):
        context = super(ListBannerTaskDetailView, self).get_context_data(**kwargs)
        context["total"] = Banner.objects.filter( materia_id= self.kwargs['pk'] ).values("tarea__tipo","cod_tarea", "observacion" ).distinct().count()
        return context

    def get_queryset(self):
        data = []
        estudiantes = Banner.objects.filter( materia_id= self.kwargs['pk'] ).values("tarea__tipo","cod_tarea", "observacion" ).distinct()
        for i in range(len(estudiantes)):
            datos = {"codigo":estudiantes[i]["cod_tarea"], "tarea":str(estudiantes[i]["cod_tarea"]) + " - " + str(estudiantes[i]["tarea__tipo"]),
                     "observacion": str(estudiantes[i]["observacion"]),
                     }
            data.append(datos)
        
        return data

#Vista que elimina las tareas
class BannerTaskDeleteView(DeleteView):
    template_name = 'banner/BannerTaskDeleteView.html'
    model = Banner

    def get_object(self, queryset=None):
        if queryset is None:
            queryset = self.get_queryset()

        materias = self.kwargs['pk']
        tarea = Banner.objects.filter( cod_tarea= self.kwargs['pk'] ).values("tarea__tipo" ).distinct()
        context = {'materias':materias, "tarea": tarea}
        return context
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["tarea"] = Banner.objects.filter( cod_tarea= self.kwargs['pk'] ).values("tarea__tipo","cod_tarea" ).distinct()
        context["codigo"] = self.kwargs['pk']
        return context
    
    
    def post(self, request, *args, **kwargs):
        estudiante = Banner.objects.filter(cod_tarea=self.kwargs['pk']  ).delete()

        return HttpResponseRedirect(reverse_lazy('student_app:list-student'))
    

# vista que genera los cargues masivos de notas
class BannerNoteMasive(View):
    
    #momstramos el form que hemos creado
    def get(self, request, *args, **kwargs):
        form = BannerFormCharge()
        pk = self.kwargs['pk']
        context = {"form": form, "pk":pk }
        return render(request, r"banner\bannerNotes.html", context)
    

class ExportNotesCsv(View):

    def post(self, request, *args, **kwargs):

        
        materia = Banner.objects.filter(
            cod_tarea=self.kwargs['pk']
        ).values("student__nombre","student__apellidos").distinct()

        wb = Workbook()
        ws1 = wb.create_sheet(index=0, title="Plantilla")

        #####3333

        for number in range(0, len(CORTE)):
            c1 = ws1.cell(row=1, column=number + 1)
            c1.value = CORTE[number] + "_" + str(corte)

        var_est = 2
        for number in materia:
            c1 = ws1.cell(row=var_est, column=1)
            c1.value = number.cod_student.codigo
            c1 = ws1.cell(row=var_est, column=2)
            c1.value = str(number.cod_student)
            c1 = ws1.cell(row=var_est, column=3)
            c1.value = str(number.materia.codigo)
            var_est += 1

        content = save_virtual_workbook(wb)
        response = HttpResponse(content)
        response['Content-Disposition'] = 'attachment; filename=cargue_notas_estudiantes.xlsx'
        response['Content-Type'] = 'application/x-xlsx'
        return response