from django.db.models import Sum, Avg
from django.urls import reverse_lazy
from django.views.generic import (TemplateView,
                                  FormView, CreateView, DeleteView, UpdateView,
                                  DetailView, ListView, View)
from django.utils import timezone
from django.http import HttpResponseRedirect, HttpResponse, FileResponse
from django.shortcuts import render
from django.utils.timezone import make_aware
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_control


from applications.Student.models import *
from applications.Programs.models import *
from applications.Programs.models import *
from applications.User.mixins import *
from .models import *
from .forms import *


import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.writer.excel import save_virtual_workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from io import BytesIO

#Vista para crear gastos
class FinanceSpendCreateview(AdminRequiredMixin,CreateView):
    model = Gastos
    form_class = SpendForm
    template_name = 'finance/create_spend.html'
    success_url = reverse_lazy('finance_app:finance-general-list-view')

    def get_context_data(self, **kwargs):
        context = super(FinanceSpendCreateview, self).get_context_data(**kwargs)
        max_consecutivo_facturasSub = FacturasSub.objects.aggregate(Max('consecutivo'))['consecutivo__max'] or 0
        max_consecutivo_gastos = Gastos.objects.aggregate(Max('consecutivo'))['consecutivo__max'] or 0
        max_consecutivo_otrosIngresos = OtroIngreso.objects.aggregate(Max('consecutivo'))['consecutivo__max'] or 0
        max_consecutivo_nominas = Nominas.objects.aggregate(Max('consecutivo'))['consecutivo__max'] or 0
        preconsecutivo = max(max_consecutivo_facturasSub, max_consecutivo_gastos, max_consecutivo_otrosIngresos, max_consecutivo_nominas) + 1
        context["preconsecutivo"] = preconsecutivo
        return context

    def post(self, request, *args, **kwargs):

        formulario = SpendForm(self.request.POST)
        if formulario.is_valid():
            formulario.save()
            mensaje = f'{self.model.__name__} registrado correctamente'
            error = "No hay error!"
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 201
            return response
        else:
            mensaje1 =[]
            mensaje1.append(
                {"error": formulario.errors}
            )
            response = JsonResponse(mensaje1, safe=False)
            response.status_code = 400
            return response
        

#vista para crear otros ingresos

class FinanceOtherIncomesCreateview(AdminRequiredMixin,CreateView):
    model = OtroIngreso
    form_class = OtherIncomesForm
    template_name = 'finance/create_OtherIncomes.html'
    success_url = reverse_lazy('finance_app:finance-general-list-view')


    def get_context_data(self, **kwargs):
        context = super(FinanceOtherIncomesCreateview, self).get_context_data(**kwargs)
        max_consecutivo_facturasSub = FacturasSub.objects.aggregate(Max('consecutivo'))['consecutivo__max'] or 0
        max_consecutivo_gastos = Gastos.objects.aggregate(Max('consecutivo'))['consecutivo__max'] or 0
        max_consecutivo_otrosIngresos = OtroIngreso.objects.aggregate(Max('consecutivo'))['consecutivo__max'] or 0
        max_consecutivo_nominas = Nominas.objects.aggregate(Max('consecutivo'))['consecutivo__max'] or 0
        preconsecutivo = max(max_consecutivo_facturasSub, max_consecutivo_gastos, max_consecutivo_otrosIngresos, max_consecutivo_nominas) + 1
        context["preconsecutivo"] = preconsecutivo
        context["pk"]=self.kwargs['pk'] 
        return context

    def post(self, request, *args, **kwargs):

        formulario = OtherIncomesForm(self.request.POST)
        if formulario.is_valid():
            formulario.save()
            mensaje = f'{self.model.__name__} registrado correctamente'
            error = "No hay error!"
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 201
            return response
        else:
            mensaje1 =[]
            mensaje1.append(
                {"error": formulario.errors}
            )
            response = JsonResponse(mensaje1, safe=False)
            response.status_code = 400
            return response
        
#vista para registrar los pagos de mensualidades de los estudiantes
class FinanceInvoiceUpdateStudent(AdminRequiredMixin,CreateView):
    template_name = 'finance/update_invoice.html'
    model = Facturas
    form_class = FacturasForm

    def get_context_data(self, **kwargs):
        context = super(FinanceInvoiceUpdateStudent, self).get_context_data(**kwargs)
        max_consecutivo_facturasSub = FacturasSub.objects.aggregate(Max('consecutivo'))['consecutivo__max'] or 0
        max_consecutivo_gastos = Gastos.objects.aggregate(Max('consecutivo'))['consecutivo__max'] or 0
        max_consecutivo_otrosIngresos = OtroIngreso.objects.aggregate(Max('consecutivo'))['consecutivo__max'] or 0
        max_consecutivo_nominas = Nominas.objects.aggregate(Max('consecutivo'))['consecutivo__max'] or 0
        preconsecutivo = max(max_consecutivo_facturasSub, max_consecutivo_gastos, max_consecutivo_otrosIngresos, max_consecutivo_nominas) + 1
        pagado = FacturasSub.objects.filter(
            facturas_id=self.kwargs['pk']).aggregate(pagados=Sum("pagado"))
        data = Facturas.objects.get(pk=self.kwargs['pk'])
        monto = data.monto
        pagado2 = pagado['pagados']
        pendiente = Facturas.objects.manejo(data, pagado2)
        datos = []
        info = {'pk':self.kwargs['pk'], "preconsecutivo":preconsecutivo, "codigo": data.codigo, "monto": f'$ {monto:,.2f}', 'pagado': f'$ {pagado2 if pagado2 else 0.0 :,.2f}',
                'pendiente': f'$ {pendiente:,.2f}', 'pendiente2': pendiente}
        datos.append(info)
        context["datos"] = datos
        return context
    
    def post(self, request, *args, **kwargs):

        #Obtenemos las variables del formulario
        mensaje1 = []
        pendiente = self.request.POST.get('pendiente')
        facturas = self.request.POST.get('facturas')
        observacion = self.request.POST.get('observacion')
        consecutivo = self.request.POST.get('consecutivo')
        pagado = self.request.POST.get('pagado')
        characters = ",0"
        pendiente = ''.join(x for x in pendiente if x not in characters)
        formulario = FacturasForm(self.request.POST)

        if formulario.is_valid():
            #Creamos la factura relacionada a la factura principal
            matricula = FacturasSub.objects.create(
                facturas=Facturas.objects.get(codigo=facturas),
                observacion=observacion,
                consecutivo=consecutivo,
                pagado=pagado,
            )
            fact_act = FacturasSub.objects.filter(facturas_id=Facturas.objects.get(
                codigo=facturas).id).aggregate(Sum('pagado'))
            monto = Facturas.objects.get(codigo=facturas).monto
            if int(fact_act['pagado__sum']) == int(monto):
                Facturas.objects.filter(codigo=facturas).update(
                    estado_id=CatalogsTypesInvoices.objects.get(estado="Pagada"))
            elif int(fact_act['pagado__sum']) > 0 and int(fact_act['pagado__sum']) < int(monto):
                Facturas.objects.filter(codigo=facturas).update(
                    estado_id=CatalogsTypesInvoices.objects.get(estado="Abono"))
            else:
                Facturas.objects.filter(codigo=facturas).update(
                    estado_id=CatalogsTypesInvoices.objects.get(estado="Pendiente"))
                
            mensaje = f'{self.model.__name__} registrado correctamente'
            error = "No hay error!"
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 201
            return response
        else:
            mensaje1 =[]
            mensaje1.append(
                {"error": formulario.errors}
            )
            response = JsonResponse(mensaje1, safe=False)
            response.status_code = 400
            return response




#vista para listar las facturas de los estudiantes, pendientes, abonadas y pagadas.
@method_decorator(cache_control(no_cache=True, must_revalidate=True, no_store=True), name='dispatch')
class FinanceInvoiceListviewStudent(AdminRequiredMixin,ListView):
    model = Facturas
    template_name = 'finance/list_invoices.html'
    context_object_name = 'invoice'

    def get_context_data(self, **kwargs):
        context = super(FinanceInvoiceListviewStudent, self).get_context_data(**kwargs)
        context["datos"] = Estudiante.objects.get(slug=self.kwargs['slug'])
        context['progreso'] = int(int(Facturas.objects.filter(user_id=Estudiante.objects.get(slug=self.kwargs['slug']).codigo, estado_id=3).count())/int(Programas.objects.get(
            id=Estudiante.objects.get(slug=self.kwargs['slug']).carrera_id
        ).cuotas+2)*100)
        return context

    def get_queryset(self):
        info = []
        datos = Facturas.objects.filter(user_id=Estudiante.objects.get(
            slug=self.kwargs['slug']).codigo).order_by("codigo")
        
        fecha_hoy = timezone.now()
        
        for i in datos:
            pagado = FacturasSub.objects.filter(
                facturas_id=i.pk).aggregate(pagados=Sum("pagado"))
            data = pagado['pagados'] if pagado['pagados'] is not None else 0
            
           
            datos_final = {'pk': i.pk, 'slug':self.kwargs['slug'], "codigo": i.codigo, 'descripcion': i.descripcion, "fecha": i.due_at,
                        'estado': "Vencida" if i.due_at < fecha_hoy and (data < i.monto or data == 0) else i.estado, 
                        'monto': f'$ {i.monto:,.0f}', 'pagado': f'$ {data if data else 0.0 :,.0f}'}

            info.append(datos_final)

        return info

#Vista para generar informes diarios y por fechas


#vista para listar ingresos, gastos y otros ingresos
@method_decorator(cache_control(no_cache=True, must_revalidate=True, no_store=True), name='dispatch')
class FinanceGeneralListView(AdminRequiredMixin,ListView):
    model = FacturasSub
    template_name = 'finance/list_income.html'
    context_object_name = 'income'
    ordering =  ['-consecutivo'  ]

    def get_queryset(self):

        data = []
        income = FacturasSub.objects.all()
        if income:
            for i in income:
                data_json = {
                    "factor":"Ingreso","codigo":i.facturas.codigo, "consecutivo": i.consecutivo, 
                    "monto": f'$ {i.pagado if i.pagado else 0.0 :,.2f}', "fecha": i.created_at, "tipo": i.facturas.descripcion
                }
                data.append(data_json)
        
        spend = Gastos.objects.all()
        if spend:
            for i in spend:
                data_json = {
                    "factor":"Gasto","codigo":i.codigo, "consecutivo": i.consecutivo, 
                    "monto": f'$ {i.monto if i.monto else 0.0 :,.2f}', "fecha": i.fecha, "tipo": i.tipo
                }
                data.append(data_json)
        
        others_income = OtroIngreso.objects.all()
        if others_income:
            for i in others_income:
                data_json = {
                    "factor":"Ingreso","codigo":i.codigo, "consecutivo": i.consecutivo, 
                    "monto": f'$ {i.monto if i.monto else 0.0 :,.2f}', "fecha": i.fecha, "tipo": i.tipo
                }
                data.append(data_json)

        nominas = Nominas.objects.all()
        if nominas:
            for i in nominas:
                data_json = {
                    "factor":"Gasto","codigo":i.codigo, "consecutivo": i.consecutivo, 
                    "monto": f'$ {i.monto if i.monto else 0.0 :,.2f}', "fecha": i.fecha, "tipo": "Nómina"
                }
                data.append(data_json)


        return data



#Listar las facturas sub teniendo en cuenta o filtrando en el id
@method_decorator(cache_control(no_cache=True, must_revalidate=True, no_store=True), name='dispatch')
class FinanceFacturasSubFilterDetailView(View):

    def get(self, request, *args, **kwargs):
        info = []
        pagados_data = FacturasSub.objects.filter(
            facturas_id=Facturas.objects.get(id=int(self.request.GET.get('info'))))

        for e in pagados_data:
            datos = {"pk": e.id, 'consecutivo': e.consecutivo, 'observacion': e.observacion, 'payed': f'$ {e.pagado:,.2f}', 'fecha': str(
                e.created_at.day) + "-" + str(e.created_at.month) + "-" + str(e.created_at.year)}
            info.append(datos)
        response = {}
        response['data'] = info
        return JsonResponse(response)


#Detalle de las facturas sub
class FinanceFacturasSubDetailView(DetailView):
    template_name = 'finance/DetailSub.html'
    model = FacturasSub


#Eliminar las facturassub
class FinanceFacturasSubDeleteView(AdminRequiredMixin,DeleteView):
    template_name = 'finance/DeleteFactSub.html'
    model = FacturasSub
    
    
    def post(self, request, *args, **kwargs):

        data_delete = self.kwargs['pk']

        restar = FacturasSub.objects.get(id=data_delete).pagado
        facturas = FacturasSub.objects.get(id=data_delete).facturas_id
        fact_act = FacturasSub.objects.filter(
            facturas_id=facturas).aggregate(Sum('pagado'))
        fact_act = float(fact_act['pagado__sum']) - float(restar)

        #Monto original de la factura
        monto = Facturas.objects.get(id=facturas).monto
        Fac_delete = FacturasSub.objects.filter(id=data_delete).delete()
        
        #verificamos que tengamos valores consignados y que no esté vencida la factura
        if fact_act > 0 and fact_act < float(monto):
            Facturas.objects.filter(id=facturas).update(
                estado_id=CatalogsTypesInvoices.objects.get(estado="Abono"))
        else:
            Facturas.objects.filter(id=facturas).update(
                estado_id=CatalogsTypesInvoices.objects.get(estado="Pendiente"))
        return HttpResponseRedirect(
            self.request.META.get("HTTP_REFERER")
        )



#Genera informe de los estudiantes candidatos a grado
class FinanceAcademicInformeView(View):

    def post(self, request, *args, **kwargs):

        #Capturamos los valores que vienen del formulario
        estudiantes = str(self.request.POST.get('concat2'))
        select = self.request.POST.get('informe12')
        estudiantes = estudiantes.split(sep=",")
        listaDataAcademic = []
        listaDataFinance = []


        #iteramos sobre las notas y las facturas de los estudiantes que solicita el usuario
        for i in estudiantes:
            infoAcademic =Banner.objects.filter(student_id=i).values(
                "student_id", "student__nombre", "student__apellidos",
                "student__codigo",'materia__materia__nombre_materia').annotate(
                    promedio_calificacion=Avg('calificacion'))
            infoFinance = Facturas.objects.filter(
                user_id=Estudiante.objects.get(id=i).codigo)

            for j in infoAcademic:
                dataAcademic = {
                    "Codigo": j['student__codigo'], "Nombre": j['student__nombre'] + " " +
                    j['student__apellidos'], "Tipo": "Académico",
                    "Materia": j['materia__materia__nombre_materia'], "Promedio": float(j['promedio_calificacion'])
                }
                listaDataAcademic.append(dataAcademic)

            for l in infoFinance:
                total = FacturasSub.objects.filter(facturas_id=l.pk).aggregate(Sum('pagado'))['pagado__sum']or 0
                dataFinance = {
                    "Codigo": l.user.codigo, "Nombre": l.user.nombres + " " +
                    l.user.apellidos, "Tipo": "Financiero",
                    "Factura": l.codigo, "Estado": l.estado.estado,
                    "Total":total,
                }
                listaDataFinance.append(dataFinance)

        #mandamos todo a dataframe para su posterior creacion en el excel
        dfAcademic = pd.DataFrame(listaDataAcademic)
        dfFinance = pd.DataFrame(listaDataFinance)

        # Creacion del archivo
        if select == "1":  # informe Academico
            wb = Workbook()
            ws = wb.create_sheet(index=0, title="Academico")

            headsList = list(dfAcademic.columns.values)

            for number in range(0, len(headsList)):
                Heads = ws.cell(row=1, column=number + 1)
                Heads.value = headsList[number]

            var_est = 2
            for number in range(0, len(dfAcademic)):

                c1 = ws.cell(row=var_est, column=1)
                c1.value = dfAcademic["Codigo"][number]
                c2 = ws.cell(row=var_est, column=2)
                c2.value = dfAcademic["Nombre"][number]
                c3 = ws.cell(row=var_est, column=3)
                c3.value = dfAcademic["Tipo"][number]
                c4 = ws.cell(row=var_est, column=4)
                c4.value = dfAcademic["Materia"][number]
                c5 = ws.cell(row=var_est, column=5)
                c5.value = dfAcademic["Promedio"][number]
                var_est += 1
            content = save_virtual_workbook(wb)
            response = HttpResponse(content)
            response['Content-Disposition'] = 'attachment; filename=informe_Academico.xlsx'
            response['Content-Type'] = 'application/x-xlsx'
            return response

        else:
            # informe financiero

            wb = Workbook()
            ws = wb.create_sheet(index=0, title="Financiero")

            headsList = list(dfFinance.columns.values)

            for number in range(0, len(headsList)):
                Heads = ws.cell(row=1, column=number + 1)
                Heads.value = headsList[number]

            var_est = 2
            for number in range(0, len(dfFinance)):

                c1 = ws.cell(row=var_est, column=1)
                c1.value = dfFinance["Codigo"][number]
                c2 = ws.cell(row=var_est, column=2)
                c2.value = dfFinance["Nombre"][number]
                c3 = ws.cell(row=var_est, column=3)
                c3.value = dfFinance["Tipo"][number]
                c4 = ws.cell(row=var_est, column=4)
                c4.value = dfFinance["Factura"][number]
                c5 = ws.cell(row=var_est, column=5)
                c5.value = dfFinance["Estado"][number]
                c6 = ws.cell(row=var_est, column=6)
                c6.value = dfFinance["Total"][number]
                var_est += 1
            content = save_virtual_workbook(wb)
            response = HttpResponse(content)
            response['Content-Disposition'] = 'attachment; filename=informe_Financiero.xlsx'
            response['Content-Type'] = 'application/x-xlsx'
            return response


#Creacion de las nóminas tanto docentes como personal administrativo
class FinanceNominaView(AdminRequiredMixin,CreateView):
    model = Nominas
    form_class = NominasForm
    template_name = 'finance/create_nomina.html'
    success_url = reverse_lazy('finance_app:finance-general-list-view')

    def get_context_data(self, **kwargs):
        context = super(FinanceNominaView, self).get_context_data(**kwargs)
        max_consecutivo_facturasSub = FacturasSub.objects.aggregate(Max('consecutivo'))['consecutivo__max'] or 0
        max_consecutivo_gastos = Gastos.objects.aggregate(Max('consecutivo'))['consecutivo__max'] or 0
        max_consecutivo_otrosIngresos = OtroIngreso.objects.aggregate(Max('consecutivo'))['consecutivo__max'] or 0
        max_consecutivo_nominas = Nominas.objects.aggregate(Max('consecutivo'))['consecutivo__max'] or 0
        preconsecutivo = max(max_consecutivo_facturasSub, max_consecutivo_gastos, max_consecutivo_otrosIngresos, max_consecutivo_nominas) + 1
        context["preconsecutivo"] = preconsecutivo
        context["pk"]=self.kwargs['pk'] 
        return context
    

    def post(self, request, *args, **kwargs):

        formulario = NominasForm(self.request.POST)
        if formulario.is_valid():
            formulario.save()
            mensaje = f'{self.model.__name__} registrado correctamente'
            error = "No hay error!"
            response = JsonResponse({"mensaje": mensaje, "error": error})
            response.status_code = 201
            return response
        else:
            mensaje1 =[]
            mensaje1.append(
                {"error": formulario.errors}
            )
            response = JsonResponse(mensaje1, safe=False)
            response.status_code = 400
            return response

@method_decorator(cache_control(no_cache=True, must_revalidate=True, no_store=True), name='dispatch')
class FinanceNominaListUserView(AdminRequiredMixin,ListView):
    model = Nominas
    template_name = 'finance/list_nominas_user.html'
    context_object_name = 'nominas'
    ordering =  ['-consecutivo'  ]

    def get_context_data(self, **kwargs):
        context = super(FinanceNominaListUserView, self).get_context_data(**kwargs)
        
        try:
            context["datos"] = User.objects.get(codigo= Docente.objects.get(slug=self.kwargs['pk']).codigo)

        except:
            context["datos"] = User.objects.get(pk= self.kwargs['pk'])
        return context
    

    def get_queryset(self):

        data = []
        
        try:
            docente = Nominas.objects.filter(user_id= Docente.objects.get(slug=self.kwargs['pk']).codigo)
            for i in docente:
                data_json = {
                    "factor":"Nómina","codigo":i.codigo, "consecutivo": i.consecutivo, 
                    "monto": f'$ {i.monto if i.monto else 0.0 :,.2f}', "fecha": i.fecha,"pk":i.pk
                }
                data.append(data_json)
        
        except:
            administrativos = Nominas.objects.filter(user_id= self.kwargs['pk'])
            for i in administrativos:
                data_json = {
                    "factor":"Nómina","codigo":i.codigo, "consecutivo": i.consecutivo, 
                    "monto": f'$ {i.monto if i.monto else 0.0 :,.2f}', "fecha": i.fecha,"pk":i.pk
                }
                data.append(data_json)
        return data



#Detalle de la nomina, editar, eliminar y generar el pdf


class FinanceNominaDetailView(DetailView):
    template_name = 'finance/detail_nominas_user.html'
    model = Nominas


class FinanceNominaDeleteView(AdminRequiredMixin,DeleteView):
    template_name = 'finance/delete_nominas_user.html'
    model = Nominas
    success_url = reverse_lazy('finance_app:finance-general-list-view')


#----Vistas que generan facturas y desprendibles de pagos

###  nombre estudiante, cantidad, concepto
class FinanceInformeTemplateView(TemplateView):
    template_name = "finance/informe_plantilla.html"

    def get(self, request):
        form = RangoFechasForm()
        return render(request, r"finance/informe_plantilla.html", {'form': form})

class GenerarNominaView(View):

    def get(self, request, pk):
        datos = Nominas.objects.get(consecutivo=pk)
        rol = str(datos.user.tipe)
        if rol == "Docente":
            datosDocentes = Docente.objects.get(codigo=Nominas.objects.get(consecutivo=pk).user_id)
            factura = {
                'consecutivo': datos.consecutivo,
                'concepto': 'Desprendible de nómina',
                'descripcion': datos.descripcion,
                'monto': f"$ {float(datos.monto):,.2f}" if datos.monto else "$ 0.00",
                'fecha': datos.fecha.strftime('%Y-%m-%d'),
                'codigo': datosDocentes.codigo,
                'nombres': datosDocentes.nombres,
                'apellidos': datosDocentes.apellidos,
                'tipodocumento': datosDocentes.tDocument,
                'documento': datosDocentes.nDocument,
                'fechaIngreso': datosDocentes.fecha_reg.strftime('%Y-%m-%d'),
            }

        else:
            datosUser = User.objects.get(codigo=Nominas.objects.get(consecutivo=pk).user_id)
            factura = {
                'consecutivo': datos.consecutivo,
                'concepto': 'Desprendible de nómina',
                'descripcion': datos.descripcion,
                'monto': f"$ {float(datos.monto):,.2f}" if datos.monto else "$ 0.00",
                'fecha': datos.fecha.strftime('%Y-%m-%d'),
                'codigo': datosUser.codigo,
                'nombres': datosUser.nombres,
                'apellidos': datosUser.apellidos,
                'fechaIngreso': datosUser.created_at.strftime('%Y-%m-%d'),
            }
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        total = factura['monto']
        styles = getSampleStyleSheet()

        logo_path = "C:/Users/crist/OneDrive/Escritorio/corporacion-platformGit/corporacion-platformGit/plataforma/static/img/home/logoazul.png"  # Asegúrate de que la ruta sea correcta
        logo = Image(logo_path, width=80, height=80) 

        title = Paragraph("<b>RECIBO DEL PAGO DE NÓMINA</b>", styles['Title'])
        table_data = [[logo, title]]

        table = Table(table_data, colWidths=[70, 400])  # Ajusta el ancho de las columnas
        table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Alinear verticalmente al centro
            ('ALIGN', (1, 0), (1, 0), 'LEFT'),  # Alinear título a la izquierda
        ]))

        # Agregar tabla al documento
        elements.append(table)
        elements.append(Spacer(1, 10)) 
        



        # Encabezado de tabla
        header_data = [
            ["Corporación Nacional De Estudios", "Siglas: C.N.E Palmira", "Dirección: Calle 31 # 20-12, colombina"],
            ["Dirección: Calle 31 # 20-12, colombina", "Ciudad: Palmira - Valle", "Teléfono: 2855359 - 3185824051"],
            ["", "", ""],
            ["Recibo número: " + str(factura['consecutivo']), "Código de Empleado: " + str(factura['codigo']),
             "Fecha de Pago: " + str(factura['fecha'])],
            ["Fecha de antiguedad: " + str(factura['fechaIngreso']), "Nombre del Empleado: " + str(factura['nombres']) +" " + str(factura['apellidos']),
            ""],

        ]
        header_table = Table(header_data, colWidths=[200, 200, 200])
        header_table.setStyle(TableStyle([
            ('SPAN', (0, 2), (2, 2)),  # Esto hará que la celda en la fila 2 abarque las 3 columnas
            ('SPAN', (0, 2), (2, 2)),
            ('LINEBELOW', (0, 2), (-1, 2), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')
        ]))
        
        elements.append(header_table)
        elements.append(Spacer(1, 20))  # Espacio después de la tabla del encabezado
        
        # Tabla de conceptos con márgenes ajustados
        data = [
            ["CONCEPTOS", "ABONOS", "DESCUENTOS"],
            [factura['descripcion'], factura['monto'], ""],
            ["Ret.Seg.Soc.(C.C.)", "", "$ 52.000"],
        ]
        
        table = Table(data, colWidths=[200, 100, 100])
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.2 * inch))

        # Total
        normal_style = styles['Normal']
        elements.append(Paragraph(f"<strong>TOTAL:</strong> {total}", normal_style))
        elements.append(Spacer(1, 0.2 * inch))

        # Pie de página
        elements.append(Paragraph("Este desprendible es un documento informativo. No se requiere firma.", normal_style))

        doc.build(elements)
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename="Desprendible_pago_"+str(factura['nombres'])+str(factura['apellidos'])+".pdf")
    


class GenerarReciboView(View):
    def get(self, request, pk):

        try:
            factura = FacturasSub.objects.get(consecutivo=pk)
            
            if factura:
                
                datosEstudiante = Estudiante.objects.get(codigo=factura.facturas.user_id)
                factura = {
                    'consecutivo': factura.consecutivo,
                    'concepto': 'Recibo de pago',
                    'descripcion': factura.facturas.descripcion + " - " + factura.observacion,
                    'monto': f"$ {float(factura.pagado):,.2f}" if factura.pagado else "$ 0.00",
                    'fecha': factura.created_at.strftime('%Y-%m-%d'),
                    'codigo': datosEstudiante.codigo,
                    'nombres': datosEstudiante.nombre,
                    'apellidos': datosEstudiante.apellidos,
                    'tipodocumento': datosEstudiante.tDocument,
                    'documento': datosEstudiante.cedula,
                    'fechaIngreso': datosEstudiante.fecha_reg.strftime('%Y-%m-%d'),
                }

                buffer = BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
                elements = []
                total = factura['monto']
                styles = getSampleStyleSheet()

                logo_path = "C:/Users/crist/OneDrive/Escritorio/corporacion-platformGit/corporacion-platformGit/plataforma/static/img/home/logoazul.png"  # Asegúrate de que la ruta sea correcta
                logo = Image(logo_path, width=80, height=80) 

                title = Paragraph("<b>RECIBO DE PAGO</b>", styles['Title'])
                table_data = [[logo, title]]

                table = Table(table_data, colWidths=[70, 400])  # Ajusta el ancho de las columnas
                table.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Alinear verticalmente al centro
                    ('ALIGN', (1, 0), (1, 0), 'LEFT'),  # Alinear título a la izquierda
                ]))

                # Agregar tabla al documento
                elements.append(table)
                elements.append(Spacer(1, 10)) 
                



                # Encabezado de tabla
                header_data = [
                    ["Corporación Nacional De Estudios", "Siglas: C.N.E Palmira", "Dirección: Calle 31 # 20-12, colombina"],
                    ["Dirección: Calle 31 # 20-12, colombina", "Ciudad: Palmira - Valle", "Teléfono: 2855359 - 3185824051"],
                    ["", "", ""],
                    ["Recibo número: " + str(factura['consecutivo']), "Documento de identidad: " + str(factura['documento']),
                    "Fecha de Pago: " + str(factura['fecha'])],
                    ["Tipo de documento: " + str(factura['tipodocumento']), "Nombre del estudiante: " + str(factura['nombres']) +" " + str(factura['apellidos']),
                    ""],

                ]
                header_table = Table(header_data, colWidths=[200, 200, 200])
                header_table.setStyle(TableStyle([
                    ('SPAN', (0, 2), (2, 2)),  # Esto hará que la celda en la fila 2 abarque las 3 columnas
                    ('SPAN', (0, 2), (2, 2)),
                    ('LINEBELOW', (0, 2), (-1, 2), 1, colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')
                ]))
                
                elements.append(header_table)
                elements.append(Spacer(1, 20))  # Espacio después de la tabla del encabezado
                
                # Tabla de conceptos con márgenes ajustados
                data = [
                    ["CONCEPTOS", "Valor", "DESCUENTOS"],
                    [factura['descripcion'], factura['monto'], ""],
                ]
                
                table = Table(data, colWidths=[200, 100, 100])
                table.setStyle(TableStyle([
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')
                ]))
                
                elements.append(table)
                elements.append(Spacer(1, 0.2 * inch))

                # Total
                normal_style = styles['Normal']
                elements.append(Paragraph(f"<strong>TOTAL:</strong> {total}", normal_style))
                elements.append(Spacer(1, 0.2 * inch))

                # Pie de página
                elements.append(Paragraph("Este desprendible es un documento informativo. No se requiere firma.", normal_style))

                doc.build(elements)
                buffer.seek(0)
                return FileResponse(buffer, as_attachment=True, filename="Desprendible_pago_"+str(factura['nombres'])+str(factura['apellidos'])+".pdf")
             
        except FacturasSub.DoesNotExist:
                try:
                    gasto = Gastos.objects.get(consecutivo=pk)
                    factura = {
                        'consecutivo': gasto.consecutivo,
                        'concepto': 'Recibo de gasto',
                        'descripcion': gasto.descripcion,
                        'monto': f"$ {float(gasto.monto):,.2f}" if gasto.monto else "$ 0.00",
                        'fecha': gasto.fecha.strftime('%Y-%m-%d'),
                        'usuario':gasto.propietario,
                        'tipo': gasto.tipo,
                    }

                    buffer = BytesIO()
                    doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
                    elements = []
                    total = factura['monto']
                    styles = getSampleStyleSheet()

                    logo_path = "C:/Users/crist/OneDrive/Escritorio/corporacion-platformGit/corporacion-platformGit/plataforma/static/img/home/logoazul.png"  # Asegúrate de que la ruta sea correcta
                    logo = Image(logo_path, width=80, height=80) 

                    title = Paragraph("<b>RECIBO DE PAGO</b>", styles['Title'])
                    table_data = [[logo, title]]

                    table = Table(table_data, colWidths=[70, 400])  # Ajusta el ancho de las columnas
                    table.setStyle(TableStyle([
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Alinear verticalmente al centro
                        ('ALIGN', (1, 0), (1, 0), 'LEFT'),  # Alinear título a la izquierda
                    ]))

                    # Agregar tabla al documento
                    elements.append(table)
                    elements.append(Spacer(1, 10)) 
                    



                    # Encabezado de tabla
                    header_data = [
                        ["Corporación Nacional De Estudios", "Siglas: C.N.E Palmira", "Dirección: Calle 31 # 20-12, colombina"],
                        ["Dirección: Calle 31 # 20-12, colombina", "Ciudad: Palmira - Valle", "Teléfono: 2855359 - 3185824051"],
                        ["", "", ""],
                        ["Recibo número: " + str(factura['consecutivo']), "Usuario: " + str(factura['usuario']),
                        "Fecha de gasto: " + str(factura['fecha'])],
                        ["Tipo de gasto: " + str(factura['tipo']), "",
                        ""],

                    ]
                    header_table = Table(header_data, colWidths=[200, 200, 200])
                    header_table.setStyle(TableStyle([
                        ('SPAN', (0, 2), (2, 2)),  # Esto hará que la celda en la fila 2 abarque las 3 columnas
                        ('SPAN', (0, 2), (2, 2)),
                        ('LINEBELOW', (0, 2), (-1, 2), 1, colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')
                    ]))
                    
                    elements.append(header_table)
                    elements.append(Spacer(1, 20))  # Espacio después de la tabla del encabezado
                    
                    # Tabla de conceptos con márgenes ajustados
                    data = [
                        ["CONCEPTOS", "Valor", "DESCUENTOS"],
                        [factura['descripcion'], factura['monto'], ""],
                    ]
                    
                    table = Table(data, colWidths=[200, 100, 100])
                    table.setStyle(TableStyle([
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')
                    ]))
                    
                    elements.append(table)
                    elements.append(Spacer(1, 0.2 * inch))

                    # Total
                    normal_style = styles['Normal']
                    elements.append(Paragraph(f"<strong>TOTAL:</strong> {total}", normal_style))
                    elements.append(Spacer(1, 0.2 * inch))

                    # Pie de página
                    elements.append(Paragraph("Este desprendible es un documento informativo. No se requiere firma.", normal_style))

                    doc.build(elements)
                    buffer.seek(0)
                    return FileResponse(buffer, as_attachment=True, filename="Desprendible_gasto_"+str(factura['usuario'])+".pdf")
                

                except Gastos.DoesNotExist:
                    try:
                        # Intentar buscar en OtroIngreso
                        factura = OtroIngreso.objects.get(consecutivo=pk)
                        datosEstudiante = Estudiante.objects.get(codigo=factura.user_id)
                        factura = {
                            'consecutivo': factura.consecutivo,
                            'concepto': 'Recibo de pago',
                            'descripcion': factura.descripcion,
                            'monto': f"$ {float(factura.monto):,.2f}" if factura.monto else "$ 0.00",
                            'fecha': factura.fecha.strftime('%Y-%m-%d'),
                            'codigo': datosEstudiante.codigo,
                            'nombres': datosEstudiante.nombre,
                            'apellidos': datosEstudiante.apellidos,
                            'tipodocumento': datosEstudiante.tDocument,
                            'documento': datosEstudiante.cedula,
                            'fechaIngreso': datosEstudiante.fecha_reg.strftime('%Y-%m-%d'),
                        }

                        buffer = BytesIO()
                        doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
                        elements = []
                        total = factura['monto']
                        styles = getSampleStyleSheet()

                        logo_path = "C:/Users/crist/OneDrive/Escritorio/corporacion-platformGit/corporacion-platformGit/plataforma/static/img/home/logoazul.png"  # Asegúrate de que la ruta sea correcta
                        logo = Image(logo_path, width=80, height=80) 

                        title = Paragraph("<b>RECIBO DE PAGO</b>", styles['Title'])
                        table_data = [[logo, title]]

                        table = Table(table_data, colWidths=[70, 400])  # Ajusta el ancho de las columnas
                        table.setStyle(TableStyle([
                            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Alinear verticalmente al centro
                            ('ALIGN', (1, 0), (1, 0), 'LEFT'),  # Alinear título a la izquierda
                        ]))

                        # Agregar tabla al documento
                        elements.append(table)
                        elements.append(Spacer(1, 10)) 
                        



                        # Encabezado de tabla
                        header_data = [
                            ["Corporación Nacional De Estudios", "Siglas: C.N.E Palmira", "Dirección: Calle 31 # 20-12, colombina"],
                            ["Dirección: Calle 31 # 20-12, colombina", "Ciudad: Palmira - Valle", "Teléfono: 2855359 - 3185824051"],
                            ["", "", ""],
                            ["Recibo número: " + str(factura['consecutivo']), "Documento de identidad: " + str(factura['documento']),
                            "Fecha de Pago: " + str(factura['fecha'])],
                            ["Tipo de documento: " + str(factura['tipodocumento']), "Nombre del estudiante: " + str(factura['nombres']) +" " + str(factura['apellidos']),
                            "Código del estudiante: " + str(factura['codigo'])],

                        ]
                        header_table = Table(header_data, colWidths=[200, 200, 200])
                        header_table.setStyle(TableStyle([
                            ('SPAN', (0, 2), (2, 2)),  # Esto hará que la celda en la fila 2 abarque las 3 columnas
                            ('SPAN', (0, 2), (2, 2)),
                            ('LINEBELOW', (0, 2), (-1, 2), 1, colors.black),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')
                        ]))
                        
                        elements.append(header_table)
                        elements.append(Spacer(1, 20))  # Espacio después de la tabla del encabezado
                        
                        # Tabla de conceptos con márgenes ajustados
                        data = [
                            ["CONCEPTOS", "Valor", "DESCUENTOS"],
                            [factura['descripcion'], factura['monto'], ""],
                        ]
                        
                        table = Table(data, colWidths=[200, 100, 100])
                        table.setStyle(TableStyle([
                            ('GRID', (0, 0), (-1, -1), 1, colors.black),
                            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold')
                        ]))
                        
                        elements.append(table)
                        elements.append(Spacer(1, 0.2 * inch))

                        # Total
                        normal_style = styles['Normal']
                        elements.append(Paragraph(f"<strong>TOTAL:</strong> {total}", normal_style))
                        elements.append(Spacer(1, 0.2 * inch))

                        # Pie de página
                        elements.append(Paragraph("Este desprendible es un documento informativo. No se requiere firma.", normal_style))

                        doc.build(elements)
                        buffer.seek(0)
                        return FileResponse(buffer, as_attachment=True, filename="Desprendible_gasto_"+str(factura['nombres'])+str(factura['apellidos'])+".pdf")

                        
                    except OtroIngreso.DoesNotExist:
                        # Si no se encuentra en ninguna tabla, devolver un error 404
                        return HttpResponse(f"Error inesperado gasto3", status=500)
        except Exception as e:
            # Manejo genérico de excepciones
            return HttpResponse(f"Error inesperado: {str(e)}", status=500) 

class GenerarInformeDiarioDiaView(View):
    def get(self, request):
        fecha_hoy =  timezone.now()
        fecha_hoy  = fecha_hoy.date()

        pagos = FacturasSub.objects.filter(created_at__date=fecha_hoy)
        gastos = Gastos.objects.filter(fecha__date=fecha_hoy)
        otrosIngresos = OtroIngreso.objects.filter(fecha__date=fecha_hoy)
        nomina = Nominas.objects.filter(fecha__date=fecha_hoy)

        resultados = [
            {"fecha":fecha_hoy,"recibo": p.consecutivo, "nombre": p.facturas.user.nombres + " " + p.facturas.user.apellidos, "tipo": "Ingreso", "concepto": p.observacion, "ingreso": p.pagado, "egreso": None}
            for p in pagos
        ] + [
            {"fecha":fecha_hoy,"recibo": g.consecutivo, "nombre": g.propietario, "tipo": "Gasto", "concepto": g.descripcion, "ingreso": None, "egreso": g.monto}
            for g in gastos
        ] + [
            {"fecha":fecha_hoy,"recibo": oi.consecutivo, "nombre": oi.user.nombres+ " " + oi.user.apellidos, "tipo": "Ingreso", "concepto": oi.descripcion, "ingreso": oi.monto, "egreso": None}
            for oi in otrosIngresos
        ] + [
            {"fecha":fecha_hoy,"recibo": n.consecutivo, "nombre": n.user.nombres+ " " + n.user.apellidos, "tipo": "Gasto", "concepto": n.descripcion, "ingreso": None, "egreso": n.monto}
            for n in nomina
        ]


        # Crear libro y hoja de cálculo
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Informe actual"

        # Encabezados
        headers = ["Fecha","Recibo", "Nombre", "Tipo", "Concepto", "Ingreso", "Egreso"]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Llenar datos
        for row_num, resultado in enumerate(resultados, 2):
            sheet.cell(row=row_num, column=1).value = resultado["fecha"]
            sheet.cell(row=row_num, column=2).value = resultado["recibo"]
            sheet.cell(row=row_num, column=3).value = resultado["nombre"]
            sheet.cell(row=row_num, column=4).value = resultado["tipo"]
            sheet.cell(row=row_num, column=5).value = resultado["concepto"]
            sheet.cell(row=row_num, column=6).value = resultado["ingreso"]
            sheet.cell(row=row_num, column=7).value = resultado["egreso"]

        # Ajustar ancho de columnas
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Obtener la letra de la columna
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            adjusted_width = max_length + 2
            sheet.column_dimensions[column].width = adjusted_width

        # Respuesta HTTP con archivo Excel
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="Informe_{fecha_hoy}.xlsx"'
        workbook.save(response)
        return response
    
class GenerarInformeDiarioRangoView(View):

    def get(self, request):
        form = RangoFechasForm()
        fecha_hoy = date.today()
        return render(request, r"finance/informe_plantilla.html", {'form': form, 'fecha_hoy': fecha_hoy})
    
    def post(self, request):
        form = RangoFechasForm(self.request.POST)
        if form.is_valid():
            fecha_inicio = form.cleaned_data['fecha_inicio']
            fecha_fin = form.cleaned_data['fecha_fin']

            # Convertir fechas a timezone-aware
            fecha_inicio = make_aware(datetime.combine(fecha_inicio, datetime.min.time()))
            fecha_fin = make_aware(datetime.combine(fecha_fin, datetime.max.time()))

            # Filtrar los datos
            pagos = FacturasSub.objects.filter(created_at__range=(fecha_inicio, fecha_fin))
            gastos = Gastos.objects.filter(fecha__range=(fecha_inicio, fecha_fin))
            otros_ingresos = OtroIngreso.objects.filter(fecha__range=(fecha_inicio, fecha_fin))
            nomina = Nominas.objects.filter(fecha__range=(fecha_inicio, fecha_fin))

            # Crear resultados
            resultados = [
                {"Fecha": p.created_at, "Recibo": p.consecutivo, 
                 "Nombre": p.facturas.user.nombres + " " + p.facturas.user.apellidos, 
                 "Tipo": "Ingreso", "Concepto": p.observacion, "Ingreso": p.pagado, "Egreso": None}
                for p in pagos
            ] + [
                {"Fecha": g.fecha, "Recibo": g.consecutivo, "Nombre": g.propietario, 
                 "Tipo": "Gasto", "Concepto": g.descripcion, "Ingreso": None, "Egreso": g.monto}
                for g in gastos
            ] + [
                {"Fecha": oi.fecha, "Recibo": oi.consecutivo, 
                 "Nombre": oi.user.nombres + " " + oi.user.apellidos, 
                 "Tipo": "Ingreso", "Concepto": oi.descripcion, "Ingreso": oi.monto, "Egreso": None}
                for oi in otros_ingresos
            ] + [
                {"Fecha": n.fecha, "Recibo": n.consecutivo, 
                 "Nombre": n.user.nombres + " " + n.user.apellidos, 
                 "Tipo": "Gasto", "Concepto": n.descripcion, "Ingreso": None, "Egreso": n.monto}
                for n in nomina
            ]

            for resultado in resultados:
                if isinstance(resultado["Fecha"], datetime) and resultado["Fecha"].tzinfo is not None:
                    resultado["Fecha"] = resultado["Fecha"].replace(tzinfo=None)

            
            wb = Workbook()
            ws = wb.active
            ws.title = "Informe"

            # Agregar encabezados
            encabezados = ["Fecha", "Recibo", "Nombre", "Tipo", "Concepto", "Ingreso", "Egreso"]
            for col_num, encabezado in enumerate(encabezados, start=1):
                ws.cell(row=1, column=col_num, value=encabezado)

            # Agregar los datos al archivo
            for row_num, fila in enumerate(resultados, start=2):
                for col_num, encabezado in enumerate(encabezados, start=1):
                    ws.cell(row=row_num, column=col_num, value=fila[encabezado])

            # Ajustar ancho de las columnas
            for col_num in range(1, len(encabezados) + 1):
                col_letter = get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = 20

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = 'attachment; filename="informe_rango_fechas.xlsx"'
            wb.save(response)
            return response
        else:
            # Volver a renderizar la plantilla con el formulario y los errores
            return render(request, r"finance/informe_plantilla.html", {"form": form})

#vista para listar las facturas de los estudiantes, pendientes, abonadas y pagadas(Solo lo ve el estudiante).
@method_decorator(cache_control(no_cache=True, must_revalidate=True, no_store=True), name='dispatch')
class FinanceInvoiceMyPayListview(StudentRequiredMixin,ListView):
    model = Facturas
    template_name = 'finance/list_invoices.html'
    context_object_name = 'invoice'

    def dispatch(self, request, *args, **kwargs):
        estudiante = Estudiante.objects.get(username=self.request.user)

        if str(estudiante.username) != str(self.request.user):
            logout(request)
            previous_url = reverse('homepage_app:logout')
            return redirect(previous_url)

        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(FinanceInvoiceMyPayListview, self).get_context_data(**kwargs)
        estudiante = Estudiante.objects.get(username=self.request.user )
        context["datos"] = Estudiante.objects.get(slug=estudiante.slug)
        context['progreso'] = int(int(Facturas.objects.filter(user_id=Estudiante.objects.get(slug=estudiante.slug).codigo, estado_id=3).count())/int(Programas.objects.get(
            id=Estudiante.objects.get(slug=estudiante.slug).carrera_id
        ).cuotas+2)*100)
        return context

    def get_queryset(self):
        info = []
        estudiante = Estudiante.objects.get(username=self.request.user )
        datos = Facturas.objects.filter(user_id=Estudiante.objects.get(
            slug=estudiante.slug).codigo).order_by("codigo")
        
        fecha_hoy = timezone.now()
        
        for i in datos:
            pagado = FacturasSub.objects.filter(
                facturas_id=i.pk).aggregate(pagados=Sum("pagado"))
            data = pagado['pagados'] if pagado['pagados'] is not None else 0
            
           
            datos_final = {'pk': i.pk, 'slug':estudiante.slug, "codigo": i.codigo, 'descripcion': i.descripcion, "fecha": i.due_at,
                        'estado': "Vencida" if i.due_at < fecha_hoy and (data < i.monto or data == 0) else i.estado, 
                        'monto': f'$ {i.monto:,.0f}', 'pagado': f'$ {data if data else 0.0 :,.0f}'}

            info.append(datos_final)

        return info